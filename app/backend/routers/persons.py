import io
import unicodedata
from dataclasses import dataclass
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..audit import log as audit_log
from ..business_scope import (
    assert_scoped_object,
    filter_query_for_business,
    get_business_by_input,
    related_business_ids,
    resolve_write_business_id,
    scoped_get,
    visible_business_id,
)
from ..deps import get_db, require_view_access
from ..models import Activity, Area, Person, PersonScheduleTemplate, ScheduleCell, User
from ..schemas import (
    PersonCreate,
    PersonImportError,
    PersonImportResult,
    PersonImportRowsRequest,
    PersonOut,
    PersonSortOrderUpdate,
    PersonUpdate,
)
from ..user_access import can_sort_person_order, is_demo_user, is_super_user

router = APIRouter(prefix="/api/persons", tags=["persons"])

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_IMPORT_BYTES = 5 * 1024 * 1024
DEFAULT_COLLAR_TYPE = "blue_collar"

HEADER_ALIASES = {
    "namn": "name",
    "name": "name",
    "person": "name",
    "personnamn": "name",
    "noman": "noman",
    "rfid": "rfid_code",
    "rfidkod": "rfid_code",
    "rfidcode": "rfid_code",
    "bricka": "rfid_code",
    "brickkod": "rfid_code",
    "badge": "rfid_code",
    "badgecode": "rfid_code",
    "arbetstyp": "collar_type",
    "personkategori": "collar_type",
    "kategori": "collar_type",
    "collartype": "collar_type",
    "workertype": "collar_type",
    "bluewhite": "collar_type",
    "bluewhitecollar": "collar_type",
    "bluewhitecolor": "collar_type",
    "hemomrade": "home_area",
    "omrade": "home_area",
    "area": "home_area",
    "homearea": "home_area",
    "verksamhet": "business",
    "business": "business",
    "huvudaktivitet": "home_activity",
    "huvudstalle": "home_activity",
    "aktivitet": "home_activity",
    "homeactivity": "home_activity",
    "activity": "home_activity",
    "sortering": "sort_order",
    "sort": "sort_order",
    "sortorder": "sort_order",
}


@dataclass(frozen=True)
class ImportPersonRow:
    row_number: int
    business: str | None
    name: str
    noman: str | None
    rfid_code: str | None
    collar_type: str
    home_area: str | None
    home_activity: str | None
    sort_order: int | None


def _required_noman(value: str | None) -> str:
    noman = (value or "").strip()
    if not noman:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="NoMan krävs")
    if len(noman) > 120:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="NoMan får vara max 120 tecken")
    return noman


def _noman_for_update(value: str | None, current_value: str | None) -> str | None:
    if value is None:
        if current_value:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="NoMan krävs")
        return None
    return _required_noman(value)


def _clean_rfid_code(value: str | None) -> str | None:
    rfid_code = (value or "").strip().upper()
    if not rfid_code:
        return None
    if len(rfid_code) > 120:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="RFID-kod far vara max 120 tecken")
    return rfid_code


def _rfid_key(value: str | None) -> str:
    return (value or "").strip().lower()


def _person_snapshot(person: Person) -> dict:
    return {
        "id": person.id,
        "business_id": person.business_id,
        "name": person.name,
        "noman": person.noman,
        "rfid_code": person.rfid_code,
        "collar_type": person.collar_type,
        "home_area_id": person.home_area_id,
        "home_activity_id": person.home_activity_id,
        "competencies": person.competencies,
        "comment": person.comment,
        "has_fixed_schedule": person.has_fixed_schedule,
        "is_active": person.is_active,
        "sort_order": person.sort_order,
    }


def _compact_key(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    without_marks = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return "".join(ch for ch in without_marks.strip().lower() if ch.isalnum())


COLLAR_TYPE_ALIASES = {
    "blue": "blue_collar",
    "bluecollar": "blue_collar",
    "bluecolor": "blue_collar",
    "bla": "blue_collar",
    "blacollar": "blue_collar",
    "blacolor": "blue_collar",
    "lager": "blue_collar",
    "produktion": "blue_collar",
    "white": "white_collar",
    "whitecollar": "white_collar",
    "whitecolor": "white_collar",
    "vit": "white_collar",
    "vitcollar": "white_collar",
    "vitcolor": "white_collar",
    "kontor": "white_collar",
}


def _parse_collar_type(value: str | None, *, row_number: int, name: str) -> tuple[str, PersonImportError | None]:
    if not value:
        return DEFAULT_COLLAR_TYPE, None
    collar_type = COLLAR_TYPE_ALIASES.get(_compact_key(value))
    if collar_type is None:
        return DEFAULT_COLLAR_TYPE, PersonImportError(
            row=row_number,
            name=name or None,
            error="Arbetstyp maste vara Blue collar eller White collar",
        )
    return collar_type, None


def _header_key(value: str | None) -> str:
    key = _compact_key(value)
    for marker in ("obligatorisk", "frivillig", "required", "optional"):
        if key.endswith(marker):
            return key[: -len(marker)]
    return key


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_mapping(headers: tuple[object, ...] | list[object] | None) -> dict[int, str]:
    if not headers:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen saknar rubrikrad")

    mapping: dict[int, str] = {}
    for index, field in enumerate(headers):
        canonical = HEADER_ALIASES.get(_header_key(_cell_text(field)))
        if canonical:
            mapping[index] = canonical

    if "name" not in set(mapping.values()):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen måste ha kolumnen namn")
    return mapping


def _parse_sort_order(value: str, *, row_number: int, name: str) -> tuple[int | None, PersonImportError | None]:
    if not value:
        return None, None
    try:
        parsed = float(value.replace(",", "."))
    except ValueError:
        return None, PersonImportError(row=row_number, name=name or None, error="Sortering måste vara ett tal")
    if not parsed.is_integer():
        return None, PersonImportError(row=row_number, name=name or None, error="Sortering måste vara ett heltal")
    return int(parsed), None


def _parse_person_import_values(raw_rows: list[tuple[int, dict[str, object]]]) -> tuple[list[ImportPersonRow], list[PersonImportError]]:
    rows: list[ImportPersonRow] = []
    errors: list[PersonImportError] = []

    for row_number, raw_values in raw_rows:
        values = {
            field: _cell_text(raw_values.get(field))
            for field in (
                "business",
                "name",
                "noman",
                "rfid_code",
                "collar_type",
                "home_area",
                "home_activity",
                "sort_order",
            )
        }
        if not any(values.values()):
            continue

        name = values.get("name", "").strip()
        if not name:
            errors.append(PersonImportError(row=row_number, error="Namn saknas"))
            continue
        if len(name) > 120:
            errors.append(PersonImportError(row=row_number, name=name, error="Namn får vara max 120 tecken"))
            continue

        noman = values.get("noman", "").strip()
        if not noman:
            errors.append(PersonImportError(row=row_number, name=name, error="NoMan saknas"))
            continue
        if len(noman) > 120:
            errors.append(PersonImportError(row=row_number, name=name, error="NoMan får vara max 120 tecken"))
            continue

        rfid_code = values.get("rfid_code", "").strip().upper() or None
        if rfid_code is not None and len(rfid_code) > 120:
            errors.append(PersonImportError(row=row_number, name=name, error="RFID-kod far vara max 120 tecken"))
            continue

        collar_type, collar_type_error = _parse_collar_type(
            values.get("collar_type", ""),
            row_number=row_number,
            name=name,
        )
        if collar_type_error is not None:
            errors.append(collar_type_error)
            continue

        sort_order, sort_error = _parse_sort_order(values.get("sort_order", ""), row_number=row_number, name=name)
        if sort_error is not None:
            errors.append(sort_error)
            continue

        rows.append(
            ImportPersonRow(
                row_number=row_number,
                business=values.get("business") or None,
                name=name,
                noman=noman,
                rfid_code=rfid_code,
                collar_type=collar_type,
                home_area=values.get("home_area") or None,
                home_activity=values.get("home_activity") or None,
                sort_order=sort_order,
            )
        )

    return rows, errors


def build_person_import_template_excel() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Personer"
    sheet.append([
        "verksamhet (frivillig)",
        "namn (obligatorisk)",
        "NoMan (obligatorisk)",
        "RFID (frivillig)",
        "arbetstyp (frivillig)",
        "hemområde (frivillig)",
        "huvudaktivitet (frivillig)",
        "sortering (frivillig)",
    ])
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 22
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 24
    sheet.column_dimensions["G"].width = 28
    sheet.column_dimensions["H"].width = 14
    sheet.freeze_panes = "A2"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def parse_person_import_excel(content: bytes) -> tuple[list[ImportPersonRow], list[PersonImportError]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen kunde inte läsas")

    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen saknar rubrikrad")

    mapping = _header_mapping(headers)
    raw_rows: list[tuple[int, dict[str, object]]] = []

    for row_number, raw_row in enumerate(rows_iter, start=2):
        raw_rows.append((
            row_number,
            {
                canonical: raw_row[index] if index < len(raw_row) else None
                for index, canonical in mapping.items()
            },
        ))

    return _parse_person_import_values(raw_rows)


def parse_person_import_rows(payload: PersonImportRowsRequest) -> tuple[list[ImportPersonRow], list[PersonImportError]]:
    raw_rows = [
        (index, row.model_dump())
        for index, row in enumerate(payload.rows, start=1)
    ]
    return _parse_person_import_values(raw_rows)


def _lookup_map(rows, *attrs: str) -> dict[str, object]:
    lookup: dict[str, object] = {}
    for row in rows:
        for attr in attrs:
            value = getattr(row, attr, None)
            key = _compact_key(str(value) if value is not None else "")
            if key:
                lookup.setdefault(key, row)
    return lookup


def _existing_person_names(db: Session, business_id: int | None) -> set[str]:
    query = db.query(Person.name)
    if business_id is not None:
        query = query.filter(Person.business_id == business_id)
    return {_compact_key(name) for (name,) in query.all()}


def _existing_person_rfid_codes(db: Session, business_id: int | None) -> set[str]:
    query = db.query(Person.rfid_code).filter(Person.rfid_code.isnot(None))
    if business_id is not None:
        query = query.filter(Person.business_id == business_id)
    return {_rfid_key(rfid_code) for (rfid_code,) in query.all() if _rfid_key(rfid_code)}


def _find_name_conflict(
    db: Session,
    name: str,
    *,
    business_id: int | None,
    exclude_person_id: int | None = None,
) -> Person | None:
    key = _compact_key(name)
    if not key:
        return None
    query = db.query(Person)
    if business_id is not None:
        query = query.filter(Person.business_id == business_id)
    for person in query.all():
        if exclude_person_id is not None and person.id == exclude_person_id:
            continue
        if _compact_key(person.name) == key:
            return person
    return None


def _find_rfid_conflict(
    db: Session,
    rfid_code: str | None,
    *,
    business_id: int | None,
    exclude_person_id: int | None = None,
) -> Person | None:
    key = _rfid_key(rfid_code)
    if not key:
        return None
    query = db.query(Person).filter(func.lower(func.trim(Person.rfid_code)) == key)
    if business_id is not None:
        query = query.filter(Person.business_id == business_id)
    if exclude_person_id is not None:
        query = query.filter(Person.id != exclude_person_id)
    return query.first()


def _next_sort_order(db: Session, business_id: int | None) -> int:
    query = db.query(func.max(Person.sort_order))
    if business_id is not None:
        query = query.filter(Person.business_id == business_id)
    return int(query.scalar() or 0) + 1


def _import_person_rows(
    rows: list[ImportPersonRow],
    errors: list[PersonImportError],
    db: Session,
    user: User,
    *,
    duplicate_error: str,
) -> PersonImportResult:
    errors = list(errors)
    default_business_id = visible_business_id(db, user)
    area_query = db.query(Area).filter(Area.is_active.is_(True))
    activity_query = db.query(Activity).filter(Activity.is_active.is_(True))
    if default_business_id is not None:
        area_query = area_query.filter(Area.business_id == default_business_id)
        activity_query = activity_query.filter(Activity.business_id == default_business_id)
    area_lookup = _lookup_map(area_query.all(), "code", "name")
    activity_lookup = _lookup_map(activity_query.all(), "code", "label")
    existing_names: set[str] = set()
    existing_names_by_business: dict[int | None, set[str]] = {}
    existing_rfid_by_business: dict[int | None, set[str]] = {}
    seen_rfid_by_business: dict[int | None, set[str]] = {}
    next_sort_by_business: dict[int | None, int] = {}
    seen_names: set[str] = set()
    created = 0

    for row in rows:
        name_key = _compact_key(row.name)
        if name_key in seen_names:
            errors.append(PersonImportError(row=row.row_number, name=row.name, error=duplicate_error))
            continue
        seen_names.add(name_key)

        if name_key in existing_names:
            errors.append(PersonImportError(row=row.row_number, name=row.name, error="Personen finns redan"))
            continue

        home_area_id = None
        area = None
        if row.home_area:
            area = area_lookup.get(_compact_key(row.home_area))
            if area is None:
                errors.append(PersonImportError(row=row.row_number, name=row.name, error="Hemområde hittades inte"))
                continue
            home_area_id = area.id

        home_activity_id = None
        activity = None
        if row.home_activity:
            activity = activity_lookup.get(_compact_key(row.home_activity))
            if activity is None:
                errors.append(PersonImportError(row=row.row_number, name=row.name, error="Huvudaktivitet hittades inte"))
                continue
            home_activity_id = activity.id
            if home_area_id is None:
                home_area_id = activity.area_id

        requested_business_id = None
        if row.business:
            business = get_business_by_input(db, row.business)
            if business is None:
                errors.append(PersonImportError(row=row.row_number, name=row.name, error="Verksamhet hittades inte"))
                continue
            requested_business_id = business.id
        business_id = resolve_write_business_id(
            db,
            user,
            requested_business_id=requested_business_id,
            related_ids=related_business_ids(area, activity),
        )
        scoped_existing_names = existing_names_by_business.setdefault(
            business_id,
            _existing_person_names(db, business_id),
        )
        if name_key in scoped_existing_names:
            errors.append(PersonImportError(row=row.row_number, name=row.name, error="Personen finns redan"))
            continue

        rfid_key = _rfid_key(row.rfid_code)
        scoped_existing_rfid = existing_rfid_by_business.setdefault(
            business_id,
            _existing_person_rfid_codes(db, business_id),
        )
        scoped_seen_rfid = seen_rfid_by_business.setdefault(business_id, set())
        if rfid_key and rfid_key in scoped_seen_rfid:
            errors.append(PersonImportError(row=row.row_number, name=row.name, error="RFID-kod ar dubblett i importen"))
            continue
        if rfid_key and rfid_key in scoped_existing_rfid:
            errors.append(PersonImportError(row=row.row_number, name=row.name, error="RFID-kod finns redan"))
            continue

        next_sort = next_sort_by_business.setdefault(business_id, _next_sort_order(db, business_id))
        sort_order = row.sort_order if row.sort_order is not None else next_sort
        if row.sort_order is None:
            next_sort_by_business[business_id] = next_sort + 1

        person = Person(
            name=row.name,
            noman=row.noman,
            rfid_code=row.rfid_code,
            collar_type=row.collar_type,
            business_id=business_id,
            home_area_id=home_area_id,
            home_activity_id=home_activity_id,
            competencies=[],
            comment=None,
            is_active=True,
            sort_order=sort_order,
        )
        db.add(person)
        db.flush()

        audit_log(
            db,
            entity_type="person",
            entity_id=person.id,
            action="import_create",
            old_value=None,
            new_value=_person_snapshot(person),
            user_id=user.id,
            business_id=person.business_id,
        )
        existing_names.add(name_key)
        scoped_existing_names.add(name_key)
        if rfid_key:
            scoped_existing_rfid.add(rfid_key)
            scoped_seen_rfid.add(rfid_key)
        created += 1

    db.commit()
    return PersonImportResult(created=created, skipped=len(errors), errors=errors)


@router.get("", response_model=list[PersonOut])
def list_persons(
    include_inactive: bool = False,
    area_id: int | None = None,
    business_id: int | None = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("persons", "view")),
) -> list[Person]:
    q = db.query(Person)
    q = filter_query_for_business(q, Person, db, user, business_id)
    if not include_inactive:
        q = q.filter(Person.is_active.is_(True))
    if area_id is not None:
        area = scoped_get(db, Area, area_id, user, detail="Område hittades inte")
        if area.is_active is not True:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Område hittades inte")
        q = q.filter(Person.home_area_id == area_id)
    return q.order_by(Person.sort_order, Person.name).all()


@router.get("/import-template")
def download_import_template(_user: User = Depends(require_view_access("personImport", "edit"))) -> Response:
    return Response(
        content=build_person_import_template_excel(),
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="personer-importmall.xlsx"'},
    )


@router.post("/import", response_model=PersonImportResult)
async def import_persons(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("personImport", "edit")),
) -> PersonImportResult:
    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Excel-filen är för stor")

    rows, errors = parse_person_import_excel(content)
    return _import_person_rows(rows, errors, db, user, duplicate_error="Dubblett i Excel-filen")


@router.post("/import-rows", response_model=PersonImportResult)
def import_person_rows(
    payload: PersonImportRowsRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("personImport", "edit")),
) -> PersonImportResult:
    rows, errors = parse_person_import_rows(payload)
    return _import_person_rows(rows, errors, db, user, duplicate_error="Dubblett i tabellen")


def _normalized_sort_slots(persons: list[Person]) -> list[int]:
    slots = sorted(int(person.sort_order or 0) for person in persons)
    normalized: list[int] = []
    previous: int | None = None
    for slot in slots:
        next_slot = slot if previous is None or slot > previous else previous + 1
        normalized.append(next_slot)
        previous = next_slot
    return normalized


def _can_sort_person_order_across_areas(user: User) -> bool:
    return is_super_user(user) or is_demo_user(user)


def _person_sort_scope_query(db: Session, user: User, requested_people: list[Person]):
    query = db.query(Person).filter(Person.is_active.is_(True))
    if _can_sort_person_order_across_areas(user):
        business_id = visible_business_id(db, user)
        if business_id is not None:
            query = query.filter(Person.business_id == business_id)
        requested_area_ids = {person.home_area_id for person in requested_people}
        if len(requested_area_ids) == 1:
            area_id = next(iter(requested_area_ids))
            query = query.filter(Person.home_area_id.is_(None) if area_id is None else Person.home_area_id == area_id)
        return query

    if user.area_id is None:
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="Användaren saknar område för personsortering")

    home_area = scoped_get(db, Area, user.area_id, user, detail="Område hittades inte")
    if home_area.is_active is not True:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Område hittades inte")
    return query.filter(Person.home_area_id == home_area.id)


@router.put("/sort-order", response_model=list[PersonOut])
def reorder_person_sort_order(
    payload: PersonSortOrderUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("personSortOrder", "edit")),
) -> list[Person]:
    if not can_sort_person_order(user):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            detail="Personsortering kräver bemanningsansvarig, admin, Super User eller demo",
        )
    ordered_ids = [int(person_id) for person_id in payload.person_ids]
    if len(ordered_ids) != len(set(ordered_ids)):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Personlistan innehåller dubbletter")

    requested_people = db.query(Person).filter(Person.id.in_(ordered_ids)).all()
    requested_by_id = {person.id: person for person in requested_people}
    if len(requested_by_id) != len(ordered_ids):
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Person hittades inte")
    for person in requested_people:
        assert_scoped_object(db, user, person, detail="Person hittades inte")
        if person.is_active is not True:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                detail="Du kan bara sortera aktiva personer",
            )
        if not _can_sort_person_order_across_areas(user) and person.home_area_id != user.area_id:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                detail="Du kan bara sortera personer med samma hemområde som ditt användarområde",
            )

    area_people_query = _person_sort_scope_query(db, user, requested_people)
    area_people = area_people_query.order_by(Person.sort_order, Person.name, Person.id).all()
    area_ids = {person.id for person in area_people}
    if set(ordered_ids) != area_ids:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Personlistan har ändrats. Läs om vyn och försök igen.")

    before_by_id = {person.id: _person_snapshot(person) for person in area_people}
    person_by_id = {person.id: person for person in area_people}
    for person_id, sort_order in zip(ordered_ids, _normalized_sort_slots(area_people), strict=True):
        person_by_id[person_id].sort_order = sort_order

    for person_id in ordered_ids:
        person = person_by_id[person_id]
        before = before_by_id[person_id]
        after = _person_snapshot(person)
        if before == after:
            continue
        audit_log(
            db,
            entity_type="person",
            entity_id=person.id,
            action="reorder",
            old_value=before,
            new_value=after,
            user_id=user.id,
            business_id=person.business_id,
        )
    db.commit()
    for person in area_people:
        db.refresh(person)
    return [person_by_id[person_id] for person_id in ordered_ids]


@router.post("", response_model=PersonOut, status_code=status.HTTP_201_CREATED)
def create_person(
    payload: PersonCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("persons", "edit")),
) -> Person:
    noman = _required_noman(payload.noman)
    rfid_code = _clean_rfid_code(payload.rfid_code)
    area = db.get(Area, payload.home_area_id) if payload.home_area_id is not None else None
    if payload.home_area_id is not None:
        assert_scoped_object(db, user, area, detail="Hemområde hittades inte")
        if area.is_active is not True:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Hemområde hittades inte")
    activity = db.get(Activity, payload.home_activity_id) if payload.home_activity_id is not None else None
    if payload.home_activity_id is not None:
        assert_scoped_object(db, user, activity, detail="Huvudaktivitet hittades inte")
        if activity.is_active is not True:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Huvudaktivitet hittades inte")
    business_id = resolve_write_business_id(
        db,
        user,
        requested_business_id=payload.business_id,
        related_ids=related_business_ids(area, activity),
    )
    existing = _find_name_conflict(db, payload.name, business_id=business_id)
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Person med samma namn finns redan")
    if _find_rfid_conflict(db, rfid_code, business_id=business_id):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="RFID-kod finns redan")
    if area is not None and area.business_id != business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Hemområde tillhör annan verksamhet")
    if activity is not None and activity.business_id != business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Huvudaktivitet tillhör annan verksamhet")

    data = payload.model_dump()
    data["noman"] = noman
    data["rfid_code"] = rfid_code
    data["collar_type"] = data.get("collar_type") or DEFAULT_COLLAR_TYPE
    data["business_id"] = business_id
    data["is_active"] = True
    person = Person(**data)
    db.add(person)
    db.flush()
    audit_log(
        db,
        entity_type="person",
        entity_id=person.id,
        action="create",
        old_value=None,
        new_value=_person_snapshot(person),
        user_id=user.id,
        business_id=person.business_id,
    )
    db.commit()
    db.refresh(person)
    return person


@router.get("/{person_id}", response_model=PersonOut)
def get_person(person_id: int, db: Session = Depends(get_db), _user: User = Depends(require_view_access("persons", "view"))) -> Person:
    return scoped_get(db, Person, person_id, _user, detail="Person hittades inte")


@router.put("/{person_id}", response_model=PersonOut)
def update_person(
    person_id: int,
    payload: PersonUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("persons", "edit")),
) -> Person:
    person = scoped_get(db, Person, person_id, user, detail="Person hittades inte")
    if payload.name is not None and _find_name_conflict(
        db,
        payload.name,
        business_id=person.business_id,
        exclude_person_id=person_id,
    ):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Person med samma namn finns redan")
    before = _person_snapshot(person)
    data = payload.model_dump(exclude_unset=True, exclude={"is_active"})
    if "noman" in data:
        data["noman"] = _noman_for_update(data["noman"], person.noman)
    if "rfid_code" in data:
        data["rfid_code"] = _clean_rfid_code(data["rfid_code"])
        if _find_rfid_conflict(db, data["rfid_code"], business_id=person.business_id, exclude_person_id=person_id):
            raise HTTPException(status.HTTP_409_CONFLICT, detail="RFID-kod finns redan")
    if "collar_type" in data:
        data["collar_type"] = data["collar_type"] or DEFAULT_COLLAR_TYPE
    if "business_id" in data and data["business_id"] != person.business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Person kan inte flyttas mellan verksamheter")
    if "home_area_id" in data and data["home_area_id"] is not None:
        area = scoped_get(db, Area, data["home_area_id"], user, detail="Hemområde hittades inte")
        if area.is_active is not True:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Hemområde hittades inte")
        if area.business_id != person.business_id:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="Hemområde tillhör annan verksamhet")
    if "home_activity_id" in data and data["home_activity_id"] is not None:
        activity = scoped_get(db, Activity, data["home_activity_id"], user, detail="Huvudaktivitet hittades inte")
        if activity.is_active is not True:
            raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Huvudaktivitet hittades inte")
        if activity.business_id != person.business_id:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="Huvudaktivitet tillhör annan verksamhet")
    for key, value in data.items():
        setattr(person, key, value)
    person.is_active = True
    audit_log(
        db,
        entity_type="person",
        entity_id=person.id,
        action="update",
        old_value=before,
        new_value=_person_snapshot(person),
        user_id=user.id,
        business_id=person.business_id,
    )
    db.commit()
    db.refresh(person)
    return person


@router.delete("/{person_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_person(
    person_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_view_access("persons", "edit")),
) -> None:
    person = scoped_get(db, Person, person_id, user, detail="Person hittades inte")
    before = _person_snapshot(person)
    db.query(ScheduleCell).filter(ScheduleCell.person_id == person_id).delete(synchronize_session=False)
    db.query(PersonScheduleTemplate).filter(PersonScheduleTemplate.person_id == person_id).delete(
        synchronize_session=False
    )
    db.delete(person)
    audit_log(
        db,
        entity_type="person",
        entity_id=person.id,
        action="delete",
        old_value=before,
        new_value=None,
        user_id=user.id,
        business_id=person.business_id,
    )
    db.commit()

import io
import re
import unicodedata
from dataclasses import dataclass
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..audit import log as audit_log
from ..business_scope import (
    assert_scoped_object,
    filter_query_for_business,
    get_business_by_input,
    related_business_ids,
    resolve_write_business_id,
    scoped_get,
    visible_business_id,
)
from ..deps import get_current_user, get_db, require_view_access
from ..models import Activity, Area, Person, ScheduleCell, User
from ..schemas import (
    ActivityCreate,
    ActivityImportError,
    ActivityImportResult,
    ActivityImportRowsRequest,
    ActivityOut,
    ActivityUpdate,
)
from ..user_access import is_super_user

router = APIRouter(prefix="/api/activities", tags=["activities"])

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_IMPORT_BYTES = 5 * 1024 * 1024

HEADER_ALIASES = {
    "etikett": "label",
    "aktivitet": "label",
    "stalle": "label",
    "ställe": "label",
    "namn": "label",
    "label": "label",
    "name": "label",
    "activity": "label",
    "omrade": "area",
    "område": "area",
    "avdelning": "area",
    "area": "area",
    "department": "area",
    "verksamhet": "business",
    "business": "business",
    "summerassom": "summary_activity",
    "summerasom": "summary_activity",
    "summeringsaktivitet": "summary_activity",
    "sammanstallning": "summary_activity",
    "sammanställning": "summary_activity",
    "huvudaktivitet": "summary_activity",
    "huvudstalle": "summary_activity",
    "huvudställe": "summary_activity",
    "summary": "summary_activity",
    "summaryactivity": "summary_activity",
    "sortering": "sort_order",
    "sort": "sort_order",
    "sortorder": "sort_order",
}


@dataclass(frozen=True)
class ImportActivityRow:
    row_number: int
    business: str | None
    label: str
    area: str | None
    summary_activity: str | None
    sort_order: int | None


def _code_part(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", normalized).strip("_")
    return normalized.upper()


def _compact_key(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    without_marks = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return "".join(ch for ch in without_marks.strip().lower() if ch.isalnum())


def _header_key(value: str | None) -> str:
    key = _compact_key(value)
    for marker in ("obligatorisk", "frivillig", "required", "optional"):
        if key.endswith(marker):
            return key[: -len(marker)]
    return key


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_mapping(headers: tuple[object, ...] | list[object] | None) -> dict[int, str]:
    if not headers:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen saknar rubrikrad")

    mapping: dict[int, str] = {}
    for index, field in enumerate(headers):
        canonical = HEADER_ALIASES.get(_header_key(_cell_text(field)))
        if canonical:
            mapping[index] = canonical

    if "label" not in set(mapping.values()):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen måste ha kolumnen etikett")
    return mapping


def _parse_sort_order(value: str, *, row_number: int, label: str) -> tuple[int | None, ActivityImportError | None]:
    if not value:
        return None, None
    try:
        parsed = float(value.replace(",", "."))
    except ValueError:
        return None, ActivityImportError(row=row_number, label=label or None, error="Sortering måste vara ett tal")
    if not parsed.is_integer():
        return None, ActivityImportError(row=row_number, label=label or None, error="Sortering måste vara ett heltal")
    return int(parsed), None


def _parse_activity_import_values(
    raw_rows: list[tuple[int, dict[str, object]]],
) -> tuple[list[ImportActivityRow], list[ActivityImportError]]:
    rows: list[ImportActivityRow] = []
    errors: list[ActivityImportError] = []

    for row_number, raw_values in raw_rows:
        values = {
            field: _cell_text(raw_values.get(field))
            for field in ("business", "label", "area", "summary_activity", "sort_order")
        }
        if not any(values.values()):
            continue

        label = values.get("label", "").strip()
        if not label:
            errors.append(ActivityImportError(row=row_number, error="Etikett saknas"))
            continue
        if len(label) > 60:
            errors.append(ActivityImportError(row=row_number, label=label, error="Etikett får vara max 60 tecken"))
            continue

        sort_order, sort_error = _parse_sort_order(values.get("sort_order", ""), row_number=row_number, label=label)
        if sort_error is not None:
            errors.append(sort_error)
            continue

        rows.append(
            ImportActivityRow(
                row_number=row_number,
                business=values.get("business") or None,
                label=label,
                area=values.get("area") or None,
                summary_activity=values.get("summary_activity") or None,
                sort_order=sort_order,
            )
        )

    return rows, errors


def build_activity_import_template_excel() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Aktiviteter"
    sheet.append(["verksamhet (frivillig)", "etikett (obligatorisk)", "område (frivillig)", "summeras som (frivillig)", "sortering (frivillig)"])
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 14
    sheet.freeze_panes = "A2"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def parse_activity_import_excel(content: bytes) -> tuple[list[ImportActivityRow], list[ActivityImportError]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen kunde inte läsas")

    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen saknar rubrikrad")

    mapping = _header_mapping(headers)
    raw_rows: list[tuple[int, dict[str, object]]] = []

    for row_number, raw_row in enumerate(rows_iter, start=2):
        raw_rows.append((
            row_number,
            {
                canonical: raw_row[index] if index < len(raw_row) else None
                for index, canonical in mapping.items()
            },
        ))

    return _parse_activity_import_values(raw_rows)


def parse_activity_import_rows(payload: ActivityImportRowsRequest) -> tuple[list[ImportActivityRow], list[ActivityImportError]]:
    raw_rows = [
        (index, row.model_dump())
        for index, row in enumerate(payload.rows, start=1)
    ]
    return _parse_activity_import_values(raw_rows)


def _activity_code_base(label: str, area: Area | None) -> str:
    label_part = _code_part(label) or "AKTIVITET"
    area_part = _code_part(area.code if area else None)
    if area_part and label_part != area_part and not label_part.startswith(f"{area_part}_"):
        return f"{area_part}_{label_part}"
    return label_part


def _unique_activity_code(db: Session, base: str, business_id: int | None) -> str:
    base = (base or "AKTIVITET")[:40].rstrip("_") or "AKTIVITET"
    candidate = base
    suffix = 2
    while db.query(Activity).filter_by(business_id=business_id, code=candidate).first():
        suffix_text = f"_{suffix}"
        candidate = f"{base[:40 - len(suffix_text)].rstrip('_')}{suffix_text}"
        suffix += 1
    return candidate


def _resolve_activity_code(db: Session, payload: ActivityCreate, admin: User, business_id: int | None) -> str:
    provided = (payload.code or "").strip()
    if provided:
        if not is_super_user(admin):
            raise HTTPException(status.HTTP_403_FORBIDDEN, detail="Endast Super User kan ange aktivitetskod")
        code = _code_part(provided)
        if not code:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Aktivitetskod saknar giltiga tecken")
        if db.query(Activity).filter_by(business_id=business_id, code=code).first():
            raise HTTPException(status.HTTP_409_CONFLICT, detail="Aktivitet med samma kod finns redan")
        return code

    area = db.get(Area, payload.area_id) if payload.area_id is not None else None
    return _unique_activity_code(db, _activity_code_base(payload.label, area), business_id)


def _activity_snapshot(activity: Activity) -> dict:
    return {
        "id": activity.id,
        "business_id": activity.business_id,
        "code": activity.code,
        "label": activity.label,
        "area_id": activity.area_id,
        "summary_activity_id": activity.summary_activity_id,
        "color": activity.color,
        "category": activity.category,
        "sort_order": activity.sort_order,
        "is_active": activity.is_active,
        "required_competency": activity.required_competency,
    }


def _lookup_map(rows, *attrs: str) -> dict[str, object]:
    lookup: dict[str, object] = {}
    for row in rows:
        for attr in attrs:
            value = getattr(row, attr, None)
            key = _compact_key(str(value) if value is not None else "")
            if key:
                lookup.setdefault(key, row)
    return lookup


def _existing_activity_labels(db: Session, business_id: int | None) -> set[str]:
    query = db.query(Activity.label)
    if business_id is not None:
        query = query.filter(Activity.business_id == business_id)
    return {_compact_key(label) for (label,) in query.all()}


def _next_sort_order(db: Session, business_id: int | None) -> int:
    query = db.query(func.max(Activity.sort_order))
    if business_id is not None:
        query = query.filter(Activity.business_id == business_id)
    return int(query.scalar() or 0) + 1


def _validate_summary_activity(
    db: Session,
    *,
    activity_id: int | None,
    summary_activity_id: int | None,
    business_id: int | None,
) -> int | None:
    if summary_activity_id is None:
        return None
    if activity_id is not None and summary_activity_id == activity_id:
        return None

    target = db.get(Activity, summary_activity_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Summeringsaktivitet hittades inte")
    if business_id is not None and target.business_id != business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Summeringsaktivitet tillhör annan verksamhet")

    if activity_id is None:
        return summary_activity_id

    visited = {activity_id}
    current = target
    while current.summary_activity_id is not None:
        if current.summary_activity_id in visited:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="Summeringskoppling skapar en loop")
        visited.add(current.id)
        current = db.get(Activity, current.summary_activity_id)
        if current is None:
            break

    return summary_activity_id


def _import_activity_rows(
    rows: list[ImportActivityRow],
    errors: list[ActivityImportError],
    db: Session,
    admin: User,
    *,
    duplicate_error: str,
) -> ActivityImportResult:
    errors = list(errors)
    default_business_id = visible_business_id(db, admin)
    area_query = db.query(Area)
    activity_query = db.query(Activity)
    if default_business_id is not None:
        area_query = area_query.filter(Area.business_id == default_business_id)
        activity_query = activity_query.filter(Activity.business_id == default_business_id)
    area_lookup = _lookup_map(area_query.all(), "code", "name")
    activity_lookup = _lookup_map(activity_query.all(), "code", "label")
    existing_labels: set[str] = set()
    existing_labels_by_business: dict[int | None, set[str]] = {}
    next_sort_by_business: dict[int | None, int] = {}
    seen_labels: set[str] = set()
    created = 0

    for row in rows:
        label_key = _compact_key(row.label)
        if label_key in seen_labels:
            errors.append(ActivityImportError(row=row.row_number, label=row.label, error=duplicate_error))
            continue
        seen_labels.add(label_key)

        if label_key in existing_labels:
            errors.append(ActivityImportError(row=row.row_number, label=row.label, error="Aktiviteten finns redan"))
            continue

        area = None
        if row.area:
            area = area_lookup.get(_compact_key(row.area))
            if area is None:
                errors.append(ActivityImportError(row=row.row_number, label=row.label, error="Område hittades inte"))
                continue

        summary_activity_id = None
        summary_activity = None
        if row.summary_activity:
            summary_activity = activity_lookup.get(_compact_key(row.summary_activity))
            if summary_activity is None:
                errors.append(
                    ActivityImportError(row=row.row_number, label=row.label, error="Summeringsaktivitet hittades inte")
                )
                continue

        requested_business_id = None
        if row.business:
            business = get_business_by_input(db, row.business)
            if business is None:
                errors.append(ActivityImportError(row=row.row_number, label=row.label, error="Verksamhet hittades inte"))
                continue
            requested_business_id = business.id
        business_id = resolve_write_business_id(
            db,
            admin,
            requested_business_id=requested_business_id,
            related_ids=related_business_ids(area, summary_activity),
        )
        scoped_existing_labels = existing_labels_by_business.setdefault(
            business_id,
            _existing_activity_labels(db, business_id),
        )
        if label_key in scoped_existing_labels:
            errors.append(ActivityImportError(row=row.row_number, label=row.label, error="Aktiviteten finns redan"))
            continue
        summary_activity_id = _validate_summary_activity(
            db,
            activity_id=None,
            summary_activity_id=summary_activity.id if summary_activity is not None else None,
            business_id=business_id,
        )

        next_sort = next_sort_by_business.setdefault(business_id, _next_sort_order(db, business_id))
        sort_order = row.sort_order if row.sort_order is not None else next_sort
        if row.sort_order is None:
            next_sort_by_business[business_id] = next_sort + 1

        code = _unique_activity_code(db, _activity_code_base(row.label, area), business_id)
        activity = Activity(
            code=code,
            label=row.label,
            business_id=business_id,
            area_id=area.id if area is not None else None,
            summary_activity_id=summary_activity_id,
            color="#ffffff",
            category="work",
            sort_order=sort_order,
            is_active=True,
            required_competency=None,
        )
        db.add(activity)
        db.flush()

        audit_log(
            db,
            entity_type="activity",
            entity_id=activity.id,
            action="import_create",
            old_value=None,
            new_value=_activity_snapshot(activity),
            user_id=admin.id,
            business_id=activity.business_id,
        )
        existing_labels.add(label_key)
        scoped_existing_labels.add(label_key)
        activity_lookup.setdefault(label_key, activity)
        activity_lookup.setdefault(_compact_key(activity.code), activity)
        created += 1

    db.commit()
    return ActivityImportResult(created=created, skipped=len(errors), errors=errors)


@router.get("", response_model=list[ActivityOut])
def list_activities(
    include_inactive: bool = False,
    business_id: int | None = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[Activity]:
    q = filter_query_for_business(db.query(Activity), Activity, db, user, business_id)
    _ = include_inactive
    return q.order_by(Activity.sort_order, Activity.label).all()


@router.get("/import-template")
def download_import_template(_admin: User = Depends(require_view_access("activityImport", "edit"))) -> Response:
    return Response(
        content=build_activity_import_template_excel(),
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="aktiviteter-importmall.xlsx"'},
    )


@router.post("/import", response_model=ActivityImportResult)
async def import_activities(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("activityImport", "edit")),
) -> ActivityImportResult:
    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Excel-filen är för stor")

    rows, errors = parse_activity_import_excel(content)
    return _import_activity_rows(rows, errors, db, admin, duplicate_error="Dubblett i Excel-filen")


@router.post("/import-rows", response_model=ActivityImportResult)
def import_activity_rows(
    payload: ActivityImportRowsRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("activityImport", "edit")),
) -> ActivityImportResult:
    rows, errors = parse_activity_import_rows(payload)
    return _import_activity_rows(rows, errors, db, admin, duplicate_error="Dubblett i tabellen")


@router.post("", response_model=ActivityOut, status_code=status.HTTP_201_CREATED)
def create_activity(
    payload: ActivityCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("activities", "edit")),
) -> Activity:
    area = db.get(Area, payload.area_id) if payload.area_id is not None else None
    if payload.area_id is not None:
        assert_scoped_object(db, admin, area, detail="Område hittades inte")
    summary_activity = db.get(Activity, payload.summary_activity_id) if payload.summary_activity_id is not None else None
    if payload.summary_activity_id is not None:
        assert_scoped_object(db, admin, summary_activity, detail="Summeringsaktivitet hittades inte")
    business_id = resolve_write_business_id(
        db,
        admin,
        requested_business_id=payload.business_id,
        related_ids=related_business_ids(area, summary_activity),
    )
    data = payload.model_dump()
    data["business_id"] = business_id
    data["code"] = _resolve_activity_code(db, payload, admin, business_id)
    data["summary_activity_id"] = _validate_summary_activity(
        db,
        activity_id=None,
        summary_activity_id=payload.summary_activity_id,
        business_id=business_id,
    )
    activity = Activity(**data)
    db.add(activity)
    db.flush()
    audit_log(
        db,
        entity_type="activity",
        entity_id=activity.id,
        action="create",
        old_value=None,
        new_value=_activity_snapshot(activity),
        user_id=admin.id,
        business_id=activity.business_id,
    )
    db.commit()
    db.refresh(activity)
    return activity


@router.put("/{activity_id}", response_model=ActivityOut)
def update_activity(
    activity_id: int,
    payload: ActivityUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("activities", "edit")),
) -> Activity:
    activity = scoped_get(db, Activity, activity_id, admin, detail="Aktivitet hittades inte")
    before = _activity_snapshot(activity)
    if payload.code is not None:
        if not is_super_user(admin):
            raise HTTPException(status.HTTP_403_FORBIDDEN, detail="Endast Super User kan ändra aktivitetskod")
        payload.code = _code_part(payload.code)
        if not payload.code:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Aktivitetskod saknar giltiga tecken")
        existing = (
            db.query(Activity)
            .filter(
                Activity.business_id == activity.business_id,
                Activity.code == payload.code,
                Activity.id != activity_id,
            )
            .first()
        )
        if existing:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="Aktivitet med samma kod finns redan")
    data = payload.model_dump(exclude_unset=True)
    if "business_id" in data and data["business_id"] != activity.business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Aktivitet kan inte flyttas mellan verksamheter")
    if "area_id" in data and data["area_id"] is not None:
        area = scoped_get(db, Area, data["area_id"], admin, detail="Område hittades inte")
        if area.business_id != activity.business_id:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="Område tillhör annan verksamhet")
    if "summary_activity_id" in data:
        if payload.summary_activity_id is not None:
            scoped_get(db, Activity, payload.summary_activity_id, admin, detail="Summeringsaktivitet hittades inte")
        data["summary_activity_id"] = _validate_summary_activity(
            db,
            activity_id=activity_id,
            summary_activity_id=payload.summary_activity_id,
            business_id=activity.business_id,
        )
    for key, value in data.items():
        setattr(activity, key, value)
    audit_log(
        db,
        entity_type="activity",
        entity_id=activity.id,
        action="update",
        old_value=before,
        new_value=_activity_snapshot(activity),
        user_id=admin.id,
        business_id=activity.business_id,
    )
    db.commit()
    db.refresh(activity)
    return activity


@router.delete("/{activity_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_activity(
    activity_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("activities", "edit")),
) -> None:
    activity = scoped_get(db, Activity, activity_id, admin, detail="Aktivitet hittades inte")
    before = _activity_snapshot(activity)
    db.query(Person).filter(Person.home_activity_id == activity_id).update(
        {Person.home_activity_id: None},
        synchronize_session=False,
    )
    db.query(Activity).filter(Activity.summary_activity_id == activity_id).update(
        {Activity.summary_activity_id: None},
        synchronize_session=False,
    )
    db.query(ScheduleCell).filter(ScheduleCell.activity_id == activity_id).update(
        {ScheduleCell.activity_id: None, ScheduleCell.empty_override: True},
        synchronize_session=False,
    )
    db.delete(activity)
    audit_log(
        db,
        entity_type="activity",
        entity_id=activity.id,
        action="delete",
        old_value=before,
        new_value=None,
        user_id=admin.id,
        business_id=activity.business_id,
    )
    db.commit()

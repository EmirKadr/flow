from __future__ import annotations

import io
import unicodedata
from dataclasses import dataclass
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from sqlalchemy.orm import Session

from .. import audit
from ..business_scope import (
    assert_scoped_object,
    filter_query_for_business,
    get_business_by_input,
    related_business_ids,
    resolve_write_business_id,
    scoped_get,
    visible_business_id,
)
from ..deps import get_db, require_view_access
from ..models import Area, User
from ..schemas import UserAdminOut, UserCreate, UserImportError, UserImportResult, UserImportRowsRequest, UserUpdate
from ..security import hash_password
from ..user_access import SUPER_USER_ROLE, can_admin, is_super_user, normalize_user_roles, primary_role, user_admin_out, user_roles

router = APIRouter(prefix="/api/users", tags=["users"])

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_IMPORT_BYTES = 5 * 1024 * 1024

HEADER_ALIASES = {
    "anvandarnamn": "username",
    "username": "username",
    "user": "username",
    "namn": "display_name",
    "name": "display_name",
    "visningsnamn": "display_name",
    "displayname": "display_name",
    "roll": "role",
    "roller": "role",
    "role": "role",
    "roles": "role",
    "avdelning": "area",
    "omrade": "area",
    "område": "area",
    "area": "area",
    "department": "area",
    "verksamhet": "business",
    "business": "business",
}

ROLE_ALIASES = {
    "admin": "admin",
    "administrator": "admin",
    "arbetsledare": "leader",
    "ledare": "leader",
    "leader": "leader",
    "bemanningsansvarig": "staffing_manager",
    "bemanningsanvarig": "staffing_manager",
    "staffingmanager": "staffing_manager",
    "staffing_manager": "staffing_manager",
    "produktionsledare": "staffing_manager",
    "produktion": "staffing_manager",
    "productionleader": "staffing_manager",
    "production_leader": "staffing_manager",
    "lagerkontorist": "warehouse_clerk",
    "lager": "warehouse_clerk",
    "warehouseclerk": "warehouse_clerk",
    "warehouse": "warehouse_clerk",
    "artikelplacerare": "article_placer",
    "artikelplacering": "article_placer",
    "articleplacer": "article_placer",
    "articleplacement": "article_placer",
    "visning": "viewer",
    "visningslage": "viewer",
    "lasare": "viewer",
    "läsare": "viewer",
    "viewer": "viewer",
}
ROLE_SPLIT_CHARS = ",;/+&"


@dataclass(frozen=True)
class ImportUserRow:
    row_number: int
    business: str | None
    username: str
    display_name: str | None
    roles: list[str]
    area_name: str | None = None


def _find_username_conflict(db: Session, username: str, *, exclude_user_id: int | None = None) -> User | None:
    query = db.query(User).filter(func.lower(User.username) == username.lower())
    if exclude_user_id is not None:
        query = query.filter(User.id != exclude_user_id)
    return query.order_by(User.id.asc()).first()


def _active_admin_count(db: Session, *, business_id: int | None = None, exclude_user_id: int | None = None) -> int:
    query = db.query(User).filter(User.is_active.is_(True))
    if business_id is not None:
        query = query.filter(User.business_id == business_id)
    if exclude_user_id is not None:
        query = query.filter(User.id != exclude_user_id)
    return sum(1 for user in query.all() if can_admin(user))


def _has_super_user_role(roles: list[str]) -> bool:
    return SUPER_USER_ROLE in {str(role or "").strip().lower() for role in roles}


def _guard_super_user_role_change(*, current_roles: list[str], new_roles: list[str], admin: User) -> None:
    if _has_super_user_role(current_roles) == _has_super_user_role(new_roles):
        return
    if is_super_user(admin):
        return
    raise HTTPException(
        status.HTTP_403_FORBIDDEN,
        detail="Endast Super User kan ändra Super User-rollen",
    )


def _user_snapshot(user: User) -> dict:
    roles = user_roles(user)
    return {
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "role": primary_role(roles),
        "roles": roles,
        "business_id": user.business_id,
        "area_id": user.area_id,
        "is_active": user.is_active,
        "must_change_password": bool(user.must_change_password or user.password_hash is None),
    }


def _compact_key(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    without_marks = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return "".join(ch for ch in without_marks.strip().lower() if ch.isalnum())


def _header_key(value: str | None) -> str:
    key = _compact_key(value)
    for marker in ("obligatorisk", "frivillig", "required", "optional"):
        if key.endswith(marker):
            return key[: -len(marker)]
    return key


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_mapping(headers: tuple[object, ...] | list[object] | None) -> dict[int, str]:
    if not headers:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen saknar rubrikrad")

    mapping: dict[int, str] = {}
    for index, field in enumerate(headers):
        canonical = HEADER_ALIASES.get(_header_key(_cell_text(field)))
        if canonical:
            mapping[index] = canonical

    missing = {"username", "display_name", "role"} - set(mapping.values())
    if missing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail="Excel-filen måste ha kolumnerna användarnamn, namn och roll",
        )
    return mapping


def _normalize_role(value: str) -> str | None:
    return ROLE_ALIASES.get(_compact_key(value))


def _normalize_roles(value: str) -> list[str] | None:
    normalized_value = value or ""
    for char in ROLE_SPLIT_CHARS:
        normalized_value = normalized_value.replace(char, ",")
    roles: list[str] = []
    for part in normalized_value.split(","):
        role = _normalize_role(part)
        if role and role not in roles:
            roles.append(role)
    return roles or None


def _roles_text(value: object) -> str:
    if isinstance(value, list):
        return ", ".join(_cell_text(item) for item in value)
    return _cell_text(value)


def _parse_user_import_values(raw_rows: list[tuple[int, dict[str, object]]]) -> tuple[list[ImportUserRow], list[UserImportError]]:
    rows: list[ImportUserRow] = []
    errors: list[UserImportError] = []

    for row_number, raw_values in raw_rows:
        values = {
            field: _cell_text(raw_values.get(field))
            for field in ("business", "username", "display_name", "role", "area")
        }
        if not values["role"]:
            values["role"] = _roles_text(raw_values.get("roles"))
        if not any(values.values()):
            continue

        username = values["username"]
        display_name = values["display_name"] or None
        role_value = values["role"]
        area_name = values.get("area") or None

        if not username:
            errors.append(UserImportError(row=row_number, error="Användarnamn saknas"))
            continue
        if len(username) > 50:
            errors.append(UserImportError(row=row_number, username=username, error="Användarnamn får vara max 50 tecken"))
            continue
        if display_name is not None and len(display_name) > 100:
            errors.append(UserImportError(row=row_number, username=username, error="Namn får vara max 100 tecken"))
            continue

        roles = _normalize_roles(role_value)
        if roles is None:
            errors.append(
                UserImportError(
                    row=row_number,
                    username=username,
                    error="Roll måste vara admin, arbetsledare, bemanningsansvarig, lagerkontorist, artikelplacerare eller visning. Flera roller kan separeras med komma.",
                )
            )
            continue

        rows.append(ImportUserRow(
            row_number=row_number,
            business=values.get("business") or None,
            username=username,
            display_name=display_name,
            roles=roles,
            area_name=area_name,
        ))

    return rows, errors


def _area_lookup(db: Session, business_id: int | None) -> dict[str, Area]:
    lookup: dict[str, Area] = {}
    query = db.query(Area)
    if business_id is not None:
        query = query.filter(Area.business_id == business_id)
    for area in query.all():
        for value in (area.code, area.name):
            key = _compact_key(value)
            if key:
                lookup[key] = area
    return lookup


def _validate_area_id(db: Session, area_id: int | None, user: User, business_id: int | None) -> None:
    if area_id is None:
        return
    area = scoped_get(db, Area, area_id, user, detail="Område hittades inte")
    if business_id is not None and area.business_id != business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Område tillhör annan verksamhet")
    if False:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Område hittades inte")


def build_user_import_template_excel() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Användare"
    sheet.append(["verksamhet (frivillig)", "användarnamn (obligatorisk)", "namn (frivillig)", "roller (obligatorisk)", "område (frivillig)"])
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 22
    sheet.freeze_panes = "A2"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def parse_user_import_excel(content: bytes) -> tuple[list[ImportUserRow], list[UserImportError]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen kunde inte läsas")

    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Excel-filen saknar rubrikrad")

    mapping = _header_mapping(headers)
    raw_rows: list[tuple[int, dict[str, object]]] = []

    for row_number, raw_row in enumerate(rows_iter, start=2):
        raw_rows.append((
            row_number,
            {
                canonical: raw_row[index] if index < len(raw_row) else None
                for index, canonical in mapping.items()
            },
        ))

    return _parse_user_import_values(raw_rows)


def parse_user_import_rows(payload: UserImportRowsRequest) -> tuple[list[ImportUserRow], list[UserImportError]]:
    raw_rows = [
        (index, row.model_dump())
        for index, row in enumerate(payload.rows, start=1)
    ]
    return _parse_user_import_values(raw_rows)


def _import_user_rows(
    rows: list[ImportUserRow],
    errors: list[UserImportError],
    db: Session,
    admin: User,
    *,
    duplicate_error: str,
) -> UserImportResult:
    errors = list(errors)
    seen: set[str] = set()
    created = 0
    default_business_id = visible_business_id(db, admin)
    areas_by_key = _area_lookup(db, default_business_id)

    for row in rows:
        username_key = row.username.lower()
        if username_key in seen:
            errors.append(UserImportError(row=row.row_number, username=row.username, error=duplicate_error))
            continue
        seen.add(username_key)

        if _find_username_conflict(db, row.username):
            errors.append(UserImportError(row=row.row_number, username=row.username, error="Användarnamnet används redan"))
            continue
        requested_business_id = None
        if row.business:
            business = get_business_by_input(db, row.business)
            if business is None:
                errors.append(UserImportError(row=row.row_number, username=row.username, error="Verksamhet hittades inte"))
                continue
            requested_business_id = business.id
        area_id = None
        area = None
        if row.area_name:
            area = areas_by_key.get(_compact_key(row.area_name))
            if not area:
                errors.append(UserImportError(row=row.row_number, username=row.username, error="Område hittades inte"))
                continue
            area_id = area.id
        business_id = resolve_write_business_id(
            db,
            admin,
            requested_business_id=requested_business_id,
            related_ids=related_business_ids(area),
        )

        user = User(
            username=row.username,
            password_hash=None,
            display_name=row.display_name,
            role=primary_role(row.roles),
            roles=row.roles,
            business_id=business_id,
            area_id=area_id,
            is_active=True,
            must_change_password=True,
        )
        db.add(user)
        db.flush()

        audit.log(
            db,
            entity_type="user",
            entity_id=user.id,
            action="import_create",
            old_value=None,
            new_value=_user_snapshot(user),
            user_id=admin.id,
            business_id=user.business_id,
        )
        created += 1

    db.commit()
    return UserImportResult(created=created, skipped=len(errors), errors=errors)


@router.get("", response_model=list[UserAdminOut])
def list_users(
    include_inactive: bool = False,
    business_id: int | None = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("users", "view")),
) -> list[UserAdminOut]:
    query = filter_query_for_business(db.query(User), User, db, admin, business_id)
    if not include_inactive:
        query = query.filter(User.is_active.is_(True))
    users = query.order_by(User.username.asc()).all()
    users.sort(key=lambda user: (0 if can_admin(user) else 1, user.username.lower()))
    return [user_admin_out(user) for user in users]


@router.get("/import-template")
def download_import_template(_admin: User = Depends(require_view_access("userImport", "edit"))) -> Response:
    return Response(
        content=build_user_import_template_excel(),
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="anvandare-importmall.xlsx"'},
    )


@router.post("/import", response_model=UserImportResult)
async def import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("userImport", "edit")),
) -> UserImportResult:
    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Excel-filen är för stor")

    rows, errors = parse_user_import_excel(content)
    return _import_user_rows(rows, errors, db, admin, duplicate_error="Dubblett i Excel-filen")


@router.post("/import-rows", response_model=UserImportResult)
def import_user_rows(
    payload: UserImportRowsRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("userImport", "edit")),
) -> UserImportResult:
    rows, errors = parse_user_import_rows(payload)
    return _import_user_rows(rows, errors, db, admin, duplicate_error="Dubblett i tabellen")


@router.post("", response_model=UserAdminOut, status_code=status.HTTP_201_CREATED)
def create_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("users", "edit")),
) -> UserAdminOut:
    if _find_username_conflict(db, payload.username):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Användarnamnet används redan")
    area = db.get(Area, payload.area_id) if payload.area_id is not None else None
    if payload.area_id is not None:
        assert_scoped_object(db, admin, area, detail="Område hittades inte")
    business_id = resolve_write_business_id(
        db,
        admin,
        requested_business_id=payload.business_id,
        related_ids=related_business_ids(area),
    )
    _validate_area_id(db, payload.area_id, admin, business_id)
    roles = normalize_user_roles(payload.roles, payload.role)
    _guard_super_user_role_change(current_roles=[], new_roles=roles, admin=admin)

    user = User(
        username=payload.username,
        password_hash=hash_password(payload.password) if payload.password else None,
        display_name=payload.display_name,
        role=primary_role(roles),
        roles=roles,
        business_id=business_id,
        area_id=payload.area_id,
        is_active=payload.is_active,
        must_change_password=payload.password is None,
    )
    db.add(user)
    db.flush()

    audit.log(
        db,
        entity_type="user",
        entity_id=user.id,
        action="create",
        old_value=None,
        new_value=_user_snapshot(user),
        user_id=admin.id,
        business_id=user.business_id,
    )

    db.commit()
    db.refresh(user)
    return user_admin_out(user)


@router.put("/{user_id}", response_model=UserAdminOut)
def update_user(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_view_access("users", "edit")),
) -> UserAdminOut:
    user = scoped_get(db, User, user_id, admin, detail="Användare hittades inte")
    if not user:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Användare hittades inte")

    if payload.username is not None and _find_username_conflict(db, payload.username, exclude_user_id=user_id):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Användarnamnet används redan")
    area = db.get(Area, payload.area_id) if payload.area_id is not None else None
    if payload.area_id is not None:
        assert_scoped_object(db, admin, area, detail="Område hittades inte")
    target_business_id = user.business_id
    if payload.business_id is not None:
        target_business_id = resolve_write_business_id(
            db,
            admin,
            requested_business_id=payload.business_id,
            related_ids=related_business_ids(area),
        )
    if area is not None and area.business_id != target_business_id:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Område tillhör annan verksamhet")

    current_roles = user_roles(user)
    if payload.roles is not None or payload.role is not None:
        new_roles = normalize_user_roles(payload.roles, payload.role or primary_role(current_roles))
    else:
        new_roles = current_roles
    _guard_super_user_role_change(current_roles=current_roles, new_roles=new_roles, admin=admin)
    new_is_active = payload.is_active if payload.is_active is not None else user.is_active
    removes_admin_access = can_admin(user) and (not can_admin(User(role=primary_role(new_roles), roles=new_roles)) or not new_is_active)
    if removes_admin_access and _active_admin_count(db, business_id=user.business_id, exclude_user_id=user.id) == 0:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail="Det måste finnas minst en aktiv administratör kvar",
        )

    before = _user_snapshot(user)
    updates = payload.model_dump(exclude_unset=True, exclude={"password", "role", "roles"})
    if "business_id" in updates:
        updates["business_id"] = target_business_id
        if "area_id" not in updates and user.area_id is not None:
            current_area = db.get(Area, user.area_id)
            if current_area is not None and current_area.business_id != target_business_id:
                updates["area_id"] = None
    for key, value in updates.items():
        setattr(user, key, value)
    if payload.roles is not None or payload.role is not None:
        user.roles = new_roles
        user.role = primary_role(new_roles)
    if payload.password is not None:
        user.password_hash = hash_password(payload.password)
        user.must_change_password = False

    audit.log(
        db,
        entity_type="user",
        entity_id=user.id,
        action="update",
        old_value=before,
        new_value=_user_snapshot(user),
        user_id=admin.id,
        business_id=user.business_id,
    )

    db.commit()
    db.refresh(user)
    return user_admin_out(user)

"""Run an interactive end-to-end exercise of the flow web app.

This is the "let an agent poke around" tool. It creates a disposable local
database, starts the app, clicks through important workflows, creates records,
edits schedule cells, verifies role restrictions, and saves screenshots plus a
JSON report.
"""
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook

from tools.terminology_contracts import assert_no_forbidden_terms_in_text, role_access_required_terms
from tools import visual_smoke


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_ROOT = ROOT / "artifacts" / "interactive"
AGENT_PASSWORD = "AgentTest123"


WEB_WORKFLOW_STEPS = (
    "login_admin",
    "create_business",
    "edit_business",
    "download_import_templates",
    "create_user",
    "edit_user",
    "import_user",
    "toggle_user_setting",
    "delete_user",
    "create_activity",
    "edit_activity",
    "delete_activity",
    "import_activity",
    "import_person",
    "create_person",
    "edit_person_inline",
    "edit_person_fields_inline",
    "edit_person_activity_inline",
    "edit_person_week_template",
    "edit_person_hourly_schedule",
    "schedule_person_activity",
    "edit_schedule_cell",
    "split_schedule_cell",
    "copy_paste_schedule_cell",
    "drag_fill_schedule_cells",
    "copy_day",
    "clear_day",
    "undo_redo",
    "overview_person_activity",
    "overview_edit",
    "history_filter",
    "role_access_click_cycle",
    "role_access_view_level",
    "role_access_edit_level",
    "role_access_none_level",
    "viewer_read_only",
    "role_access_guards",
)


@dataclass
class StepResult:
    name: str
    ok: bool
    screenshot: str | None = None
    detail: str = ""


class InteractiveRun:
    def __init__(self, page, base_url: str, output_dir: Path, run_id: str):
        self.page = page
        self.base_url = base_url.rstrip("/")
        self.output_dir = output_dir
        self.run_id = run_id
        self.results: list[StepResult] = []
        self.agent_user = f"agent_user_{run_id}"
        self.agent_user_imported = f"agent_import_user_{run_id}"
        self.agent_person = f"Agent Person {run_id}"
        self.agent_person_updated = f"Agent Person Redigerad {run_id}"
        self.agent_person_imported = f"Agent Importperson {run_id}"
        self.agent_activity = f"Agent Aktivitet {run_id}"
        self.agent_activity_updated = f"Agent Aktivitet Redigerad {run_id}"
        self.agent_activity_imported = f"Agent Importaktivitet {run_id}"
        self.agent_business_code = f"AGENT{run_id}"
        self.agent_business_name = f"Agent Verksamhet {run_id}"
        self.agent_business_updated = f"Agent Verksamhet Redigerad {run_id}"

    def url(self, path: str) -> str:
        return self.base_url + path

    def screenshot(self, name: str) -> str:
        path = self.output_dir / f"{name}.png"
        path.parent.mkdir(parents=True, exist_ok=True)
        self.page.screenshot(path=str(path), full_page=True)
        return str(path.relative_to(self.output_dir))

    def record(self, name: str, *, screenshot: str | None = None, detail: str = "") -> None:
        self.results.append(StepResult(name=name, ok=True, screenshot=screenshot, detail=detail))

    def click_and_expect_api(self, selector: str, url_part: str, method: str) -> None:
        with self.page.expect_response(
            lambda response: url_part in response.url and response.request.method == method,
            timeout=15000,
        ) as response_info:
            self.page.click(selector)
        response = response_info.value
        if not response.ok:
            body = response.text()
            self.screenshot(f"error-{method.lower()}-{url_part.strip('/').replace('/', '-')}")
            raise RuntimeError(f"{method} {url_part} failed with {response.status}: {body}")

    def goto(self, path: str, selector: str) -> None:
        self.page.goto(self.url(path), wait_until="networkidle")
        self.page.wait_for_selector(selector, timeout=15000)

    def set_area_focus(self, value: str, selector: str) -> None:
        self.page.evaluate(
            """(value) => {
                if (window.writeAreaFocus) {
                    window.writeAreaFocus(value);
                } else {
                    localStorage.setItem('flow-area-focus', value);
                    window.dispatchEvent(new CustomEvent('flow:areaFocusChanged', { detail: { value } }));
                }
            }""",
            value,
        )
        self.page.wait_for_selector(selector, timeout=15000)
        self.page.wait_for_timeout(700)

    def login(self, username: str, password: str) -> None:
        self.goto("/login.html", "#login-form")
        self.page.fill("#username", username)
        self.page.fill("#password", password)
        self.page.click("button.primary")
        self.page.wait_for_url("**/index.html", timeout=15000)
        self.page.wait_for_selector("#scheduleTable", timeout=15000)

    def logout(self) -> None:
        self.page.click("#logout-link")
        self.page.wait_for_url("**/login.html", timeout=15000)

    def wait_for_text(self, text: str) -> None:
        self.page.get_by_text(text, exact=False).first.wait_for(state="attached", timeout=15000)

    def set_color(self, selector: str, value: str) -> None:
        self.page.eval_on_selector(
            selector,
            """(el, value) => {
                el.value = value;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }""",
            value,
        )

    def dismiss_dialogs(self) -> None:
        self.page.on("dialog", lambda dialog: dialog.accept())

    def run(self) -> list[StepResult]:
        self.dismiss_dialogs()
        self.login("admin", "admin123")
        self.record("login_admin", screenshot=self.screenshot("01-admin-login"))

        self.create_business()
        self.edit_business()
        self.download_import_templates()
        self.create_user()
        self.edit_user()
        self.import_user()
        self.toggle_user_setting()
        self.delete_user()
        self.create_activity()
        self.edit_activity()
        self.delete_activity()
        self.import_activity()
        self.import_person()
        self.create_person()
        self.edit_person_inline()
        self.edit_person_fields_inline()
        self.edit_person_week_template()
        self.edit_schedule()
        self.exercise_overview()
        self.exercise_history()
        self.exercise_role_access_settings()
        self.exercise_viewer_and_guards()
        return self.results

    def create_user(self) -> None:
        self.goto("/anvandare.html", "#users-body")
        self.page.click("#new-user")
        self.page.wait_for_selector("#m-username", timeout=15000)
        self.page.fill("#m-username", self.agent_user)
        self.page.fill("#m-display-name", "Agent Test User")
        self.page.locator('input[name="m-role"][value="leader"]').set_checked(True)
        self.page.locator("#m-area").select_option(label="Mestergruppen")
        self.page.fill("#m-password", AGENT_PASSWORD)
        before = self.screenshot("02-create-user-modal")
        self.click_and_expect_api("#m-save", "/api/users", "POST")
        self.wait_for_text(self.agent_user)
        self.record("create_user", screenshot=before, detail=self.agent_user)
        self.record("create_user_saved", screenshot=self.screenshot("03-user-created"))

    def create_business(self) -> None:
        self.goto("/verksamheter.html", "#businesses-body")
        self.page.click("#new-business")
        self.page.wait_for_selector("#m-code", timeout=15000)
        self.page.fill("#m-code", self.agent_business_code)
        self.page.fill("#m-name", self.agent_business_name)
        self.page.fill("#m-sort", "90")
        modal = self.screenshot("02-create-business-modal")
        self.click_and_expect_api("#save", "/api/businesses", "POST")
        self.wait_for_text(self.agent_business_code)
        self.record("create_business", screenshot=modal, detail=self.agent_business_code)
        self.record("create_business_saved", screenshot=self.screenshot("02-business-created"))

    def edit_business(self) -> None:
        row = self.page.locator("#businesses-body tr", has_text=self.agent_business_code)
        row.locator("button[data-edit]").click()
        self.page.wait_for_selector("#m-name", timeout=15000)
        self.page.fill("#m-name", self.agent_business_updated)
        modal = self.screenshot("02b-edit-business-modal")
        self.click_and_expect_api("#save", "/api/businesses", "PUT")
        self.wait_for_text(self.agent_business_updated)
        self.record("edit_business", screenshot=modal)
        self.record("edit_business_saved", screenshot=self.screenshot("02c-business-edited"))

    def download_import_templates(self) -> None:
        downloads = (
            ("/personer.html", "#download-person-template", "personer-importmall.xlsx"),
            ("/aktiviteter.html", "#download-activity-template", "aktiviteter-importmall.xlsx"),
            ("/anvandare.html", "#download-user-template", "anvandare-importmall.xlsx"),
        )
        for path, selector, expected_name in downloads:
            self.goto(path, selector)
            with self.page.expect_download(timeout=15000) as download_info:
                self.page.click(selector)
            download = download_info.value
            if download.suggested_filename != expected_name:
                raise AssertionError(f"Expected {expected_name}, got {download.suggested_filename}")
        self.record("download_import_templates", screenshot=self.screenshot("02-import-template-downloads"))

    def edit_user(self) -> None:
        row = self.page.locator("#users-body tr", has_text=self.agent_user)
        row.locator("button[data-edit]").click()
        self.page.wait_for_selector("#m-display-name", timeout=15000)
        self.page.fill("#m-display-name", "Agent Test User Redigerad")
        self.page.locator('input[name="m-role"][value="leader"]').set_checked(False)
        self.page.locator('input[name="m-role"][value="viewer"]').set_checked(True)
        self.page.locator("#m-area").select_option(label="Granngården")
        modal = self.screenshot("04-edit-user-modal")
        self.page.click("#m-save")
        self.wait_for_text("Agent Test User Redigerad")
        self.record("edit_user", screenshot=modal)
        self.record("edit_user_saved", screenshot=self.screenshot("05-user-edited"))

    def import_user(self) -> None:
        self.goto("/anvandare.html", "#users-body")
        xlsx_path = self.output_dir / f"user-import-{self.run_id}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["anvandarnamn", "namn", "roller", "omrade"])
        sheet.append([self.agent_user_imported, "Agent Import User", "Visning", "Mestergruppen"])
        workbook.save(xlsx_path)

        self.page.set_input_files("#user-import-file", str(xlsx_path))
        self.wait_for_text(self.agent_user_imported)
        self.record("import_user", screenshot=self.screenshot("05b-user-imported"))

    def toggle_user_setting(self) -> None:
        checkbox = self.page.locator("#lock-foreign-schedule-cells")
        initial = checkbox.is_checked()
        checkbox.set_checked(not initial)
        self.page.wait_for_timeout(500)
        checkbox.set_checked(initial)
        self.page.wait_for_timeout(500)
        self.record("toggle_user_setting", screenshot=self.screenshot("06-user-setting-toggled"))

    def delete_user(self) -> None:
        self.goto("/anvandare.html", "#users-body")
        row = self.page.locator("#users-body tr", has_text=self.agent_user_imported)
        with self.page.expect_response(
            lambda response: "/api/users/" in response.url and response.request.method == "DELETE",
            timeout=15000,
        ) as response_info:
            row.locator("button[data-delete]").click()
        response = response_info.value
        if not response.ok:
            raise RuntimeError(f"DELETE user failed with {response.status}: {response.text()}")
        row.wait_for(state="detached", timeout=15000)
        self.record("delete_user", screenshot=self.screenshot("07-user-deleted"))

    def create_activity(self) -> None:
        self.goto("/aktiviteter.html", "#acts-body")
        self.page.click("#new-act")
        self.page.wait_for_selector("#m-label", timeout=15000)
        self.page.fill("#m-label", self.agent_activity)
        self.page.locator("#m-area").select_option(label="Granngården")
        self.set_color("#m-color", "#34d399")
        self.page.select_option("#m-cat", "work")
        self.page.fill("#m-sort", "998")
        modal = self.screenshot("06-create-activity-modal")
        self.page.click("#m-save")
        self.wait_for_text(self.agent_activity)
        self.record("create_activity", screenshot=modal)
        self.record("create_activity_saved", screenshot=self.screenshot("07-activity-created"))

    def edit_activity(self) -> None:
        row = self.page.locator("#acts-body tr", has_text=self.agent_activity)
        row.locator("button[data-edit]").click()
        self.page.wait_for_selector("#m-label", timeout=15000)
        self.page.fill("#m-label", self.agent_activity_updated)
        self.set_color("#m-color", "#60a5fa")
        modal = self.screenshot("08-edit-activity-modal")
        self.page.click("#m-save")
        self.wait_for_text(self.agent_activity_updated)
        self.record("edit_activity", screenshot=modal)
        self.record("edit_activity_saved", screenshot=self.screenshot("09-activity-edited"))

    def delete_activity(self) -> None:
        row = self.page.locator("#acts-body tr", has_text=self.agent_activity_updated)
        row.locator("button[data-delete]").click()
        self.page.wait_for_timeout(700)
        self.record("delete_activity", screenshot=self.screenshot("10-activity-deleted"))

    def import_activity(self) -> None:
        self.goto("/aktiviteter.html", "#acts-body")
        xlsx_path = self.output_dir / f"activity-import-{self.run_id}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["etikett", "område", "summeras som", "kategori", "färg", "sortering"])
        sheet.append([self.agent_activity_imported, "Mestergruppen", None, "arbete", "#dbeafe", 91])
        workbook.save(xlsx_path)

        self.page.set_input_files("#activity-import-file", str(xlsx_path))
        self.wait_for_text(self.agent_activity_imported)
        self.record("import_activity", screenshot=self.screenshot("11-activity-imported"))

    def import_person(self) -> None:
        self.goto("/personer.html", "#persons-body")
        xlsx_path = self.output_dir / f"person-import-{self.run_id}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["namn", "hemomrade", "huvudaktivitet", "sortering"])
        sheet.append([self.agent_person_imported, "Mestergruppen", "MG Plock", 994])
        workbook.save(xlsx_path)

        self.page.set_input_files("#person-import-file", str(xlsx_path))
        self.wait_for_text(self.agent_person_imported)
        self.record("import_person", screenshot=self.screenshot("11b-person-imported"))

    def create_person(self) -> None:
        self.goto("/personer.html", "#persons-body")
        self.page.click("#new-person")
        self.page.wait_for_selector("#m-name", timeout=15000)
        self.page.fill("#m-name", self.agent_person)
        self.page.locator("#m-area").select_option(label="Mestergruppen")
        self.page.locator("#m-activity").select_option(label="MG VM")
        self.page.fill("#m-sort", "997")
        modal = self.screenshot("10-create-person-modal")
        self.page.click("#m-save")
        self.wait_for_text(self.agent_person)
        self.record("create_person", screenshot=modal)
        self.record("create_person_saved", screenshot=self.screenshot("11-person-created"))

    def edit_person_inline(self) -> None:
        row = self.page.locator("#persons-body tr", has_text=self.agent_person)
        row.locator("td").nth(0).click()
        self.page.wait_for_selector("#persons-body input.inline-input", timeout=15000)
        self.page.fill("#persons-body input.inline-input", self.agent_person_updated)
        self.page.keyboard.press("Enter")
        self.wait_for_text(self.agent_person_updated)
        self.record("edit_person_inline", screenshot=self.screenshot("12-person-inline-edited"))

    def edit_person_fields_inline(self) -> None:
        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("td").nth(1).click()
        self.page.wait_for_selector("#persons-body select.inline-input", timeout=15000)
        self.page.locator("#persons-body select.inline-input").select_option(label="Autostore")
        self.page.wait_for_timeout(700)

        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("td").nth(1).click()
        self.page.wait_for_selector("#persons-body select.inline-input", timeout=15000)
        self.page.locator("#persons-body select.inline-input").select_option(label="Mestergruppen")
        self.page.wait_for_timeout(700)

        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("td").nth(3).click()
        self.page.wait_for_selector("#persons-body input.inline-input", timeout=15000)
        self.page.fill("#persons-body input.inline-input", "996")
        self.page.keyboard.press("Enter")
        self.page.wait_for_timeout(700)

        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("td").nth(2).click()
        self.page.wait_for_selector("#persons-body select.inline-input", timeout=15000)
        self.page.locator("#persons-body select.inline-input").select_option(label="MG Plock")
        self.page.wait_for_timeout(700)
        self.record("edit_person_activity_inline", screenshot=self.screenshot("13b-person-activity-inline-edited"))

        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("td").nth(2).click()
        self.page.wait_for_selector("#persons-body select.inline-input", timeout=15000)
        self.page.locator("#persons-body select.inline-input").select_option(label="MG VM")
        self.page.wait_for_timeout(700)

        self.record("edit_person_fields_inline", screenshot=self.screenshot("13-person-fields-inline-edited"))

    def edit_person_week_template(self) -> None:
        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("button[data-schedule]").click()
        self.page.wait_for_selector(".modal-backdrop .modal", timeout=15000)
        first_row = self.page.locator(".modal-backdrop tr[data-weekday='1']")
        first_row.locator(".m-from").select_option("8")
        first_row.locator(".m-to").select_option("17")
        self.page.locator(".modal-backdrop tr[data-weekday='6'] .m-off").check()
        self.page.locator(".modal-backdrop tr[data-weekday='7'] .m-off").check()
        modal = self.screenshot("13-week-template-modal")
        self.page.click("#sch-save")
        self.page.wait_for_selector(".modal-backdrop", state="detached", timeout=15000)
        self.record("edit_person_week_template", screenshot=modal)

        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("button[data-schedule]").click()
        self.page.wait_for_selector("#sch-hourly", timeout=15000)
        self.page.check("#sch-hourly")
        hourly_modal = self.screenshot("14-week-template-hourly")
        self.page.click("#sch-save")
        self.page.wait_for_selector(".modal-backdrop", state="detached", timeout=15000)
        self.record("edit_person_hourly_schedule", screenshot=hourly_modal)

        row = self.page.locator("#persons-body tr", has_text=self.agent_person_updated)
        row.locator("button[data-schedule]").click()
        self.page.wait_for_selector("#sch-hourly", timeout=15000)
        self.page.uncheck("#sch-hourly")
        self.page.click("#sch-default")
        self.page.click("#sch-save")
        self.page.wait_for_selector(".modal-backdrop", state="detached", timeout=15000)

    def schedule_row(self):
        return self.page.locator("#scheduleBody tr", has_text=self.agent_person_updated)

    def schedule_cell(self, hour: int):
        return self.schedule_row().locator(f"td[data-hour='{hour}']")

    def selected_schedule_label(self, hour: int) -> str:
        return self.schedule_cell(hour).locator("select").first.evaluate(
            "select => select.options[select.selectedIndex]?.textContent?.trim() || ''"
        )

    def drag_schedule_cell(self, source_hour: int, target_hour: int) -> None:
        source = self.schedule_cell(source_hour)
        target = self.schedule_cell(target_hour)
        source.scroll_into_view_if_needed()
        target.scroll_into_view_if_needed()
        source_handle = source.element_handle()
        target_handle = target.element_handle()
        if not source_handle or not target_handle:
            raise AssertionError("Could not locate schedule cells for drag fill")
        self.page.evaluate(
            """([source, target]) => {
                const sourceBox = source.getBoundingClientRect();
                const targetBox = target.getBoundingClientRect();
                const sourceX = sourceBox.left + 4;
                const sourceY = sourceBox.top + 4;
                const targetX = targetBox.left + targetBox.width / 2;
                const targetY = targetBox.top + targetBox.height / 2;
                source.dispatchEvent(new MouseEvent('mousedown', {
                    bubbles: true,
                    button: 0,
                    clientX: sourceX,
                    clientY: sourceY,
                }));
                document.dispatchEvent(new MouseEvent('mousemove', {
                    bubbles: true,
                    clientX: targetX,
                    clientY: targetY,
                }));
                document.dispatchEvent(new MouseEvent('mouseup', {
                    bubbles: true,
                    button: 0,
                    clientX: targetX,
                    clientY: targetY,
                }));
            }""",
            [source_handle, target_handle],
        )
        self.page.wait_for_timeout(1000)

    def edit_schedule(self) -> None:
        self.goto("/index.html", "#scheduleTable")
        self.set_area_focus("MG", "#scheduleBody tr")
        self.page.fill("#nameFilter", self.agent_person_updated)
        self.wait_for_text(self.agent_person_updated)

        self.schedule_cell(8).locator("select").first.select_option(label="MG Plock")
        self.page.wait_for_timeout(700)
        if self.selected_schedule_label(8) != "MG Plock":
            raise AssertionError("Schedule activity was not applied to the person row")
        self.record("schedule_person_activity", screenshot=self.screenshot("14-schedule-person-activity"))
        self.record("edit_schedule_cell", screenshot=self.screenshot("14-schedule-cell-edited"))

        self.drag_schedule_cell(8, 7)
        if self.selected_schedule_label(7) != "MG Plock":
            raise AssertionError("Drag fill did not copy the expected activity")
        self.record("drag_fill_schedule_cells", screenshot=self.screenshot("14b-schedule-drag-fill"))

        self.schedule_cell(9).dblclick()
        self.page.wait_for_selector(
            f"#scheduleBody tr:has-text('{self.agent_person_updated}') td[data-hour='9'][data-split='1']",
            timeout=15000,
        )
        self.schedule_cell(9).locator("select").nth(0).select_option(label="MG VM")
        self.page.wait_for_timeout(300)
        self.schedule_cell(9).locator("select").nth(1).select_option(label="MG Stöd")
        self.page.wait_for_timeout(700)
        self.record("split_schedule_cell", screenshot=self.screenshot("15-schedule-cell-split"))

        self.schedule_cell(8).click(position={"x": 8, "y": 8})
        self.page.keyboard.press("Control+C")
        self.schedule_cell(10).click(position={"x": 8, "y": 8})
        self.page.keyboard.press("Control+V")
        self.page.wait_for_timeout(700)
        if self.selected_schedule_label(10) != "MG Plock":
            raise AssertionError("Keyboard copy/paste did not copy the expected activity")
        self.record("copy_paste_schedule_cell", screenshot=self.screenshot("16-schedule-copy-paste"))

        self.page.click("#copyBtn")
        self.page.wait_for_selector("#cp-td", timeout=15000)
        self.page.select_option("#cp-td", "5")
        self.page.check("#cp-ow")
        modal = self.screenshot("17-copy-day-modal-filled")
        self.page.click("#cp-go")
        self.page.wait_for_selector(".modal-backdrop", state="detached", timeout=15000)
        self.page.wait_for_timeout(700)
        self.record("copy_day", screenshot=modal)
        self.record("copy_day_done", screenshot=self.screenshot("18-copy-day-done"))

        self.page.click("#clearBtn")
        self.page.wait_for_timeout(700)
        self.record("clear_day", screenshot=self.screenshot("19-clear-day-done"))

        self.page.click("#undoBtn")
        self.page.wait_for_timeout(700)
        undo_shot = self.screenshot("20-undo-after-clear")
        self.page.click("#redoBtn")
        self.page.wait_for_timeout(700)
        self.record("undo_redo", screenshot=undo_shot)
        self.record("redo_done", screenshot=self.screenshot("21-redo-after-clear"))

    def exercise_overview(self) -> None:
        self.goto("/overblick.html", "#overviewTable")
        self.set_area_focus("MG", "#overviewBody tr")
        self.page.fill("#nameFilter", self.agent_person_updated)
        self.wait_for_text(self.agent_person_updated)
        first_day_select = self.schedule_like_overview_select()
        first_day_select.select_option(label="MG Plock")
        self.page.wait_for_timeout(700)
        selected = first_day_select.evaluate(
            "select => select.options[select.selectedIndex]?.textContent?.trim() || ''"
        )
        if selected != "MG Plock":
            raise AssertionError("Overview activity was not applied to the person row")
        self.record("overview_person_activity", screenshot=self.screenshot("22-overview-person-activity"))
        self.record("overview_edit", screenshot=self.screenshot("22-overview-day-edited"))
        self.page.select_option("#viewMode", "month")
        self.page.wait_for_timeout(700)
        self.record("overview_month", screenshot=self.screenshot("23-overview-month-after-edit"))

    def schedule_like_overview_select(self):
        return self.page.locator("#overviewBody tr", has_text=self.agent_person_updated).locator("td.day select").first

    def exercise_history(self) -> None:
        self.goto("/historik.html", "#auditBody")
        self.page.select_option("#periodSelect", "all")
        self.page.fill("#actionFilter", "create")
        self.page.click("#refreshAuditBtn")
        self.page.wait_for_timeout(700)
        self.record("history_filter", screenshot=self.screenshot("24-history-filtered"))

    def open_role_access_modal(self) -> None:
        self.goto("/anvandare.html", "#users-body")
        self.page.click("#role-view-access")
        self.page.wait_for_selector("#role-access-table .role-access-toggle", timeout=15000)

    def assert_role_access_modal_uses_canonical_labels(self) -> None:
        text = self.page.locator(".role-access-modal").inner_text(timeout=15000)
        for expected in role_access_required_terms():
            if expected not in text:
                raise AssertionError(f"Role access modal is missing {expected}")
        assert_no_forbidden_terms_in_text(text, context="role access modal")

    def role_access_toggle(self, role: str, view_id: str):
        return self.page.locator(f'.role-access-toggle[data-role="{role}"][data-view="{view_id}"]')

    def assert_role_access_state(self, role: str, view_id: str, level: str, label: str) -> None:
        toggle = self.role_access_toggle(role, view_id)
        toggle.wait_for(state="visible", timeout=15000)
        actual_level = toggle.get_attribute("data-level")
        actual_label = toggle.inner_text().strip()
        if actual_level != level or actual_label != label:
            raise AssertionError(
                f"Expected {role}/{view_id} to be {level} ({label}), got {actual_level} ({actual_label})"
            )

    def set_role_access_level(self, role: str, view_id: str, level: str) -> None:
        toggle = self.role_access_toggle(role, view_id)
        toggle.wait_for(state="visible", timeout=15000)
        for _ in range(4):
            if toggle.get_attribute("data-level") == level:
                return
            toggle.click()
            self.page.wait_for_timeout(100)
        actual = toggle.get_attribute("data-level")
        raise AssertionError(f"Could not set {role}/{view_id} to {level}; current level is {actual}")

    def save_role_access_modal(self) -> None:
        self.click_and_expect_api("#role-access-save", "/api/settings/role-access", "PUT")
        self.page.wait_for_selector(".modal-backdrop", state="detached", timeout=15000)

    def restore_role_access_defaults(self) -> None:
        self.open_role_access_modal()
        self.page.click("#role-access-defaults")
        self.save_role_access_modal()
        self.record("role_access_defaults_restored", screenshot=self.screenshot("25-role-access-defaults-restored"))

    def exercise_role_access_settings(self) -> None:
        self.open_role_access_modal()
        self.assert_role_access_modal_uses_canonical_labels()
        self.page.click("#role-access-defaults")
        self.assert_role_access_state("viewer", "persons", "none", "Ingen")
        self.role_access_toggle("viewer", "persons").click()
        self.assert_role_access_state("viewer", "persons", "view", "Visa")
        self.role_access_toggle("viewer", "persons").click()
        self.assert_role_access_state("viewer", "persons", "edit", "Redigera")
        self.role_access_toggle("viewer", "persons").click()
        self.assert_role_access_state("viewer", "persons", "none", "Ingen")
        self.record("role_access_click_cycle", screenshot=self.screenshot("25-role-access-click-cycle"))

        self.set_role_access_level("viewer", "persons", "view")
        self.save_role_access_modal()
        self.logout()
        self.login("visual_viewer", visual_smoke.VISUAL_PASSWORD)
        self.page.goto(self.url("/personer.html"), wait_until="networkidle")
        self.page.wait_for_url("**/personer.html", timeout=15000)
        self.page.wait_for_selector("#persons-body tr", timeout=15000)
        self.page.wait_for_function("document.querySelector('#new-person')?.hidden === true", timeout=15000)
        if self.page.locator("#new-person").is_visible():
            raise AssertionError("Viewer with view access should not see the new-person button")
        if self.page.locator("#persons-body button[data-delete]").count() != 0:
            raise AssertionError("Viewer with view access should not see delete buttons")
        if self.page.locator("#persons-body td.editable").count() != 0:
            raise AssertionError("Viewer with view access should not get editable person cells")
        self.record("role_access_view_level", screenshot=self.screenshot("26-role-access-view-level"))

        self.logout()
        self.login("admin", "admin123")
        self.open_role_access_modal()
        self.set_role_access_level("viewer", "persons", "edit")
        self.save_role_access_modal()
        self.logout()
        self.login("visual_viewer", visual_smoke.VISUAL_PASSWORD)
        self.page.goto(self.url("/personer.html"), wait_until="networkidle")
        self.page.wait_for_url("**/personer.html", timeout=15000)
        self.page.wait_for_selector("#persons-body tr", timeout=15000)
        self.page.wait_for_function("document.querySelector('#new-person')?.hidden === false", timeout=15000)
        if not self.page.locator("#new-person").is_visible():
            raise AssertionError("Viewer with edit access should see the new-person button")
        if self.page.locator("#persons-body button[data-delete]").count() == 0:
            raise AssertionError("Viewer with edit access should see delete buttons")
        if self.page.locator("#persons-body td.editable").count() == 0:
            raise AssertionError("Viewer with edit access should get editable person cells")
        self.record("role_access_edit_level", screenshot=self.screenshot("27-role-access-edit-level"))

        self.logout()
        self.login("admin", "admin123")
        self.open_role_access_modal()
        self.set_role_access_level("viewer", "persons", "none")
        self.save_role_access_modal()
        self.logout()
        self.login("visual_viewer", visual_smoke.VISUAL_PASSWORD)
        self.page.goto(self.url("/personer.html"), wait_until="networkidle")
        self.page.wait_for_selector("#scheduleTable", timeout=15000)
        if self.page.url.endswith("/personer.html"):
            raise AssertionError("Viewer with no persons access should be redirected away from persons")
        self.record("role_access_none_level", screenshot=self.screenshot("28-role-access-none-level"))

        self.logout()
        self.login("admin", "admin123")
        self.restore_role_access_defaults()

    def exercise_viewer_and_guards(self) -> None:
        self.logout()
        self.login("visual_viewer", visual_smoke.VISUAL_PASSWORD)
        self.set_area_focus("MG", "#scheduleBody tr")
        if not self.page.locator("#copyBtn").is_disabled():
            raise AssertionError("Viewer should not be able to copy schedule days")
        if self.page.locator("#scheduleBody select:not(:disabled)").count() != 0:
            raise AssertionError("Viewer should not have editable schedule selects")
        self.record("viewer_read_only", screenshot=self.screenshot("25-viewer-read-only"))

        for path, name in (
            ("/personer.html", "viewer_blocked_personer"),
            ("/aktiviteter.html", "viewer_blocked_aktiviteter"),
            ("/anvandare.html", "viewer_blocked_anvandare"),
            ("/historik.html", "viewer_blocked_historik"),
        ):
            self.page.goto(self.url(path), wait_until="networkidle")
            self.page.wait_for_selector("#scheduleTable", timeout=15000)
            self.record(name, screenshot=self.screenshot(f"26-{name}"))

        self.logout()
        self.login("visual_leader", visual_smoke.VISUAL_PASSWORD)
        for path, name in (
            ("/anvandare.html", "leader_blocked_anvandare"),
            ("/historik.html", "leader_blocked_historik"),
        ):
            self.page.goto(self.url(path), wait_until="networkidle")
            self.page.wait_for_selector("#scheduleTable", timeout=15000)
            self.record(name, screenshot=self.screenshot(f"27-{name}"))
        self.record("role_access_guards", detail="viewer and leader blocked pages verified")


def _load_playwright():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError(
            "Playwright saknas. Kör: python -m pip install -r requirements-dev.txt "
            "och sedan: python -m playwright install chromium"
        ) from exc
    return sync_playwright


def run_web_interactive(*, base_url: str, output_dir: Path, headful: bool = False) -> list[StepResult]:
    sync_playwright = _load_playwright()
    run_id = datetime.now().strftime("%H%M%S")
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=not headful)
        context = browser.new_context(
            viewport={"width": 1440, "height": 1000},
            device_scale_factor=1,
            locale="sv-SE",
        )
        try:
            page = context.new_page()
            runner = InteractiveRun(page, base_url, output_dir, run_id)
            return runner.run()
        finally:
            context.close()
            browser.close()


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base-url", help="Use an already running server instead of starting a disposable local one.")
    parser.add_argument("--output", type=Path, help="Output directory for screenshots and report.")
    parser.add_argument("--headful", action="store_true", help="Show the browser while running the workflow.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    output_dir = args.output or (DEFAULT_OUTPUT_ROOT / datetime.now().strftime("%Y%m%d-%H%M%S"))
    output_dir.mkdir(parents=True, exist_ok=True)

    server = None
    base_url = args.base_url
    try:
        if not base_url:
            base_url, server = visual_smoke.start_local_server(output_dir)
        results = run_web_interactive(base_url=base_url, output_dir=output_dir, headful=args.headful)
    finally:
        if server:
            server.close()

    report = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "base_url": base_url,
        "expected_steps": WEB_WORKFLOW_STEPS,
        "results": [result.__dict__ for result in results],
    }
    report_path = output_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Interactive E2E completed: {len(results)} steps")
    print(f"Output written to {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

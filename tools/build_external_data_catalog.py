"""Build the private external data catalog used by Hämta data.

The generated file contains view and column structure only. It must not contain
API URLs, tokens or credentials, and it is intentionally gitignored.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_VIEWS = ROOT / "private-data" / "external_views.xlsx"
DEFAULT_COLUMNS = ROOT / "private-data" / "external_columns.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "external_data_catalog.json"


def _header_map(row: tuple) -> dict[str, int]:
    return {str(value or "").strip(): index for index, value in enumerate(row)}


def build_catalog(views_path: Path, columns_path: Path) -> dict:
    views_wb = load_workbook(views_path, read_only=True, data_only=True)
    views_ws = views_wb.active
    view_rows = views_ws.iter_rows(values_only=True)
    view_header = _header_map(next(view_rows))

    views: dict[str, dict] = {}
    for row in view_rows:
        view_id = str(row[view_header["id"]] or "").strip()
        if not view_id:
            continue
        views[view_id] = {
            "id": view_id,
            "label_en": str(row[view_header["label_en"]] or "").strip(),
            "label_sv": str(row[view_header["label_sv"]] or "").strip(),
            "columns": [],
        }

    columns_wb = load_workbook(columns_path, read_only=True, data_only=True)
    columns_ws = columns_wb.active
    column_rows = columns_ws.iter_rows(values_only=True)
    column_header = _header_map(next(column_rows))

    for row in column_rows:
        view_id = str(row[column_header["view_id"]] or "").strip()
        column_id = str(row[column_header["column_id"]] or "").strip()
        if not view_id or not column_id:
            continue
        view = views.setdefault(
            view_id,
            {"id": view_id, "label_en": "", "label_sv": "", "columns": []},
        )
        view["columns"].append(
            {
                "id": column_id,
                "order": int(row[column_header["column_order"]] or len(view["columns"]) + 1),
                "label_en": str(row[column_header["column_label_en"]] or "").strip(),
                "label_sv": str(row[column_header["column_label_sv"]] or "").strip(),
            }
        )

    result = {"version": 1, "views": sorted(views.values(), key=lambda item: item["id"])}
    for view in result["views"]:
        view["columns"].sort(key=lambda item: item["order"])
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--views", type=Path, default=DEFAULT_VIEWS)
    parser.add_argument("--columns", type=Path, default=DEFAULT_COLUMNS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    catalog = build_catalog(args.views, args.columns)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(catalog, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Skrev {len(catalog['views'])} vyer till {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

from __future__ import annotations

import csv
import os
from pathlib import Path


NAME_HINTS = {
    "v_ask_customer_order_details_all": "orders",
    "v_ask_item_summary_stock_automation": "automation",
    "v_ask_order_overview": "overview",
    "v_ask_dispatch_pallet": "dispatch",
    "v_ask_booking_putaway": "wms_booking",
    "v_ask_article_buffertpallet": "buffer",
    "v_ask_article_bufferpallet": "buffer",
    "v_ask_trans_log": "wms_trans",
    "v_ask_pick_log_full": "wms_pick",
    "item_option": "item",
    "kampanjplock": "campaign",
    "campaign": "campaign",
    "prognos idag": "prognos",
    "prognos": "prognos",
    "not_putaway": "not_putaway",
    "not putaway": "not_putaway",
    "ej_inlag": "not_putaway",
    "ej inlag": "not_putaway",
    "ejinlag": "not_putaway",
}

BUFFER_NAME_HINTS = (
    "buffertpall",
    "buffertpallet",
    "buffert_pall",
    "bufferpall",
    "bufferpallet",
    "buffer_pallet",
)


def _decode_sample(path: Path) -> str:
    data = path.read_bytes()[:128 * 1024]
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _csv_columns(path: Path) -> list[str]:
    sample = _decode_sample(path)
    if not sample.strip():
        return []
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(sample.splitlines(), dialect)
    try:
        header = next(reader)
    except StopIteration:
        return []
    if len(header) == 1 and "\t" in header[0]:
        header = header[0].split("\t")
    return [str(column).strip().lower().replace("\ufeff", "") for column in header]


def _xlsx_columns(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        sheet = workbook.active
        best: list[str] = []
        for row in sheet.iter_rows(max_row=10, values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if len(values) > len(best):
                best = values
        return [value.lower().replace("\ufeff", "") for value in best]
    finally:
        workbook.close()


def _has_any(cols: list[str], candidates: tuple[str, ...]) -> bool:
    return any(col in candidates for col in cols)


def _has_part(cols: list[str], needle: str) -> bool:
    return any(needle in col for col in cols)


def detect_file_type(path: str | os.PathLike[str]) -> str | None:
    source = Path(path)
    base_name = source.name.lower()
    for hint, file_type in NAME_HINTS.items():
        if hint in base_name:
            return file_type
    if any(hint in base_name for hint in BUFFER_NAME_HINTS):
        return "buffer"

    suffix = source.suffix.lower()
    cols = _xlsx_columns(source) if suffix in {".xlsx", ".xlsm", ".xls"} else _csv_columns(source)
    if not cols:
        return None

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        has_product_code = _has_any(cols, ("artikelnummer", "artikel", "sku", "product code", "produktkod"))
        has_quantity = _has_part(cols, "antal") or _has_any(cols, ("qty", "quantity"))
        if has_product_code and (_has_part(cols, "kampanj") or _has_part(cols, "projicerat")) and has_quantity:
            return "campaign"
        if cols == ["artikelnummer", "antal styck"] or ("artikelnummer" in cols and "antal styck" in cols):
            return "campaign"
        if has_product_code and has_quantity:
            return "prognos"
        return None

    has_art = _has_any(cols, ("artikel", "artikelnummer", "artnr", "art.nr", "sku", "article"))
    has_qty = _has_any(cols, ("bestallt", "beställt", "antal", "qty", "quantity", "order qty", "antal styck"))
    has_ord = _has_any(cols, ("ordernr", "order nr", "order number", "kund", "kundnr", "order id"))
    has_rad = _has_any(cols, ("radnr", "rad nr", "line id", "rad", "struktur", "radsnr"))
    if has_art and has_qty and (has_ord or has_rad):
        return "orders"

    has_lagerplats = _has_part(cols, "lagerplats") or _has_any(cols, ("plats", "location", "bin"))
    has_pallid = _has_any(cols, ("pallid", "pall id", "id", "sscc", "etikett", "batch"))
    has_status = _has_any(cols, ("status",))
    has_inkop = _has_part(cols, "inkopsnr") or _has_part(cols, "inköpsnr")
    has_mottaget = _has_part(cols, "mottaget")
    has_pallnr = _has_part(cols, "pall nr") or _has_part(cols, "pallnr")
    has_till = any(col == "till" or col.endswith(" till") or col.startswith("till ") for col in cols)
    has_fran = _has_part(cols, "fran") or _has_part(cols, "från")
    has_plockat = _has_part(cols, "plockat")

    if has_inkop and (has_pallnr or has_pallid) and not has_mottaget and not has_plockat:
        return "wms_booking"
    if has_lagerplats and has_pallid and has_inkop:
        return "buffer"
    if has_pallid and has_till and has_fran:
        return "wms_trans"
    if has_pallid and has_plockat and has_ord:
        return "wms_pick"
    if has_art and has_qty and has_lagerplats:
        return "buffer"

    buffer_marker_count = sum(
        1 for flag in (has_lagerplats, has_pallid, has_status, has_inkop, has_mottaget, has_pallnr) if flag
    )
    if has_art and (has_qty or has_pallid) and buffer_marker_count >= 2:
        return "buffer"

    has_pack = _has_part(cols, "pack klass") or _has_part(cols, "staplingsbar")
    has_plockpall = _has_part(cols, "plockpall")
    has_dispatch_order = _has_any(cols, ("ordernr", "order nr", "order number", "ordernummer"))
    has_dispatch_ship = any(
        "sandnings" in col or "sändnings" in col or "sandnr" in col
        for col in cols
    )
    if has_pack:
        if has_plockpall and has_dispatch_order and has_dispatch_ship:
            return "dispatch"
        return "item"

    has_ordernr = _has_any(cols, ("ordernr", "order nr", "order number"))
    has_orderdatum = _has_part(cols, "orderdatum")
    has_ordertyp = _has_part(cols, "ordertyp")
    if has_ordernr and has_orderdatum and has_dispatch_ship and has_ordertyp:
        return "overview"
    if has_plockpall and has_dispatch_order and has_dispatch_ship:
        return "dispatch"

    has_robot = _has_any(cols, ("robot",))
    has_saldo = _has_part(cols, "saldo autoplock") or _has_any(cols, ("plocksaldo", "plock saldo"))
    if has_art and (has_robot or has_saldo):
        return "automation"
    return None

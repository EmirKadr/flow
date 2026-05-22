"""Local CLI for warehouse Bearbeta/Dela flows.

This CLI runs the same flow handlers as the FastAPI allocation bridge, but
without browser, server session, cookies or IndexedDB. It is meant for
regression runs, parity checks against the old Allokera CLI, and quick local
debugging.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from . import detect
from . import flows
from .native_tables import SimpleTable, is_simple_table


DEFAULT_OUT_ROOT = Path("warehouse-cli-out")


def _slug(value: str, fallback: str = "output") -> str:
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or fallback)).strip("._-")
    return safe or fallback


def _flow_by_id(flow_id: str) -> dict:
    try:
        return flows.FLOW_BY_ID[flow_id]
    except KeyError as exc:
        raise SystemExit(f"Okant flode: {flow_id}") from exc


def _public_flow(flow: dict) -> dict:
    return {key: value for key, value in flow.items() if key != "handler"}


def _json_default(value: object) -> object:
    if isinstance(value, Path):
        return str(value)
    return str(value)


def _print_json(payload: object) -> None:
    print(json.dumps(payload, indent=2, ensure_ascii=False, default=_json_default))


def _cell(value: object) -> str:
    try:
        import pandas as pd  # type: ignore
    except Exception:  # noqa: BLE001
        pd = None

    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        return str(int(value)) if value.is_integer() else f"{value:g}"
    if pd is not None and isinstance(value, pd.Timestamp):
        return "" if pd.isna(value) else value.isoformat(sep=" ")
    text = str(value)
    return "" if text.lower() in ("nan", "nat", "none") else text


def _table_columns(table: object) -> list[str]:
    if is_simple_table(table):
        return [str(column) for column in table.columns]
    return [str(column) for column in table.columns]  # pandas-like


def _table_rows(table: object) -> list[list[str]]:
    if is_simple_table(table):
        return [[_cell(value) for value in row] for row in table.rows]
    return [[_cell(value) for value in row] for row in table.itertuples(index=False, name=None)]


def _table_len(table: object) -> int:
    return int(len(table))


def _write_table_csv(table: object, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(_table_columns(table))
        writer.writerows(_table_rows(table))


def _safe_sheet_name(value: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]+", " ", str(value or "Sheet1"))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()[:31] or "Sheet1"
    candidate = cleaned
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = (cleaned[: 31 - len(suffix)] + suffix).strip()
        counter += 1
    used.add(candidate)
    return candidate


def _write_workbook(tables: list[tuple[str, str, object]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("Openpyxl saknas for XLSX-export.") from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used: set[str] = set()
    if not tables:
        workbook.create_sheet("Resultat")
    for key, label, table in tables:
        sheet = workbook.create_sheet(_safe_sheet_name(label or key, used))
        sheet.append(_table_columns(table))
        for row in _table_rows(table):
            sheet.append(row)
    workbook.save(path)


def _write_outputs(
    *,
    flow_id: str,
    result: dict,
    out_dir: Path,
    output_format: str,
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    tables = list(result.get("tables") or [])
    outputs: dict[str, Any] = {}

    if output_format in {"csv", "both"}:
        for key, label, table in tables:
            target = out_dir / f"{_slug(key)}.csv"
            _write_table_csv(table, target)
            outputs.setdefault(key, {})["csv"] = str(target)

    if output_format in {"xlsx", "both"}:
        target = out_dir / f"{_slug(flow_id)}.xlsx"
        _write_workbook(tables, target)
        outputs["workbook"] = str(target)

    text = result.get("text")
    if text:
        target = out_dir / "text.txt"
        target.write_text(str(text), encoding="utf-8")
        outputs["text"] = str(target)

    log_lines = result.get("log") or []
    if log_lines:
        target = out_dir / "log.txt"
        target.write_text("\n".join(str(line) for line in log_lines), encoding="utf-8")
        outputs["log"] = str(target)

    return outputs


def _summary_payload(flow_id: str, result: dict, outputs: dict[str, Any]) -> dict[str, Any]:
    tables = [
        {
            "key": key,
            "label": label,
            "rows": _table_len(table),
            "columns": _table_columns(table),
            "outputs": outputs.get(key, {}),
        }
        for key, label, table in (result.get("tables") or [])
    ]
    return {
        "flow_id": flow_id,
        "summary": result.get("summary", {}),
        "display_summary": result.get("display_summary"),
        "tables": tables,
        "text": bool(result.get("text")),
        "log_lines": len(result.get("log") or []),
        "outputs": outputs,
    }


def _default_out_dir(flow_id: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return DEFAULT_OUT_ROOT / f"{_slug(flow_id)}-{stamp}"


def _load_scenario(path: Path) -> tuple[str, dict[str, Path], dict[str, str], str | None]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    flow_id = str(payload.get("flow") or payload.get("flow_id") or "").strip()
    if not flow_id:
        raise SystemExit("Scenario saknar flow/flow_id.")
    base = path.parent
    files = {
        str(key): _resolve_path(str(value), base)
        for key, value in (payload.get("files") or {}).items()
        if str(value).strip()
    }
    params = {str(key): str(value) for key, value in (payload.get("params") or {}).items() if str(value) != ""}
    return flow_id, files, params, payload.get("format")


def _resolve_path(value: str, base: Path | None = None) -> Path:
    path = Path(value)
    if not path.is_absolute() and base is not None:
        path = base / path
    return path


def _validate_payload(flow: dict, files: dict[str, Path], params: dict[str, str]) -> list[str]:
    errors: list[str] = []
    for input_def in flow.get("inputs") or []:
        key = str(input_def["key"])
        input_type = input_def.get("type")
        if input_type == "file":
            path = files.get(key)
            if input_def.get("required") and path is None:
                errors.append(f"Saknar fil: {key}")
            if path is not None and not path.exists():
                errors.append(f"Filen finns inte for {key}: {path}")
        elif input_def.get("required") and not str(params.get(key, "")).strip():
            errors.append(f"Saknar parameter: {key}")
    return errors


def _apply_auto_files(flow: dict, files: dict[str, Path], auto_files: list[str] | None) -> None:
    if not auto_files:
        return
    file_inputs = [item for item in (flow.get("inputs") or []) if item.get("type") == "file"]
    for value in auto_files:
        path = Path(value)
        if not path.exists():
            raise SystemExit(f"Filen finns inte: {path}")
        file_type = detect.detect_file_type(path)
        if not file_type:
            raise SystemExit(f"Kunde inte identifiera filtyp: {path}")
        matches = [
            str(input_def["key"])
            for input_def in file_inputs
            if file_type in (input_def.get("detect") or [])
        ]
        if not matches:
            raise SystemExit(f"Filtypen {file_type} passar inte flodet {flow['id']}: {path}")
        if len(matches) > 1:
            raise SystemExit(
                f"Filtypen {file_type} matchar flera inputs i {flow['id']}: {', '.join(matches)}. "
                "Ange filen explicit med --file key=path."
            )
        key = matches[0]
        if key in files:
            raise SystemExit(f"Input {key} har redan en fil. Ange bara en fil per input.")
        files[key] = path


def _run_flow(
    *,
    flow_id: str,
    files: dict[str, Path],
    params: dict[str, str],
    out_dir: Path,
    output_format: str,
) -> dict[str, Any]:
    flow = _flow_by_id(flow_id)
    errors = _validate_payload(flow, files, params)
    if errors:
        raise SystemExit("\n".join(errors))
    result = flow["handler"]({key: Path(path) for key, path in files.items()}, dict(params))
    outputs = _write_outputs(flow_id=flow_id, result=result, out_dir=out_dir, output_format=output_format)
    payload = _summary_payload(flow_id, result, outputs)
    (out_dir / "summary.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return payload


def _files_params_from_dynamic_args(flow: dict, args: argparse.Namespace) -> tuple[dict[str, Path], dict[str, str]]:
    files: dict[str, Path] = {}
    params: dict[str, str] = {}
    for input_def in flow.get("inputs") or []:
        key = str(input_def["key"])
        value = getattr(args, f"input_{key}", None)
        if input_def.get("type") == "file":
            if value:
                files[key] = Path(value)
        elif value is not None and str(value) != "":
            params[key] = str(value)
    return files, params


def _parse_key_values(items: list[str] | None, label: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in items or []:
        if "=" not in item:
            raise SystemExit(f"Forvantade {label}=varde, fick: {item}")
        key, value = item.split("=", 1)
        result[key] = value
    return result


def _run_dynamic_command(args: argparse.Namespace) -> int:
    flow = _flow_by_id(args.flow_id)
    files, params = _files_params_from_dynamic_args(flow, args)
    _apply_auto_files(flow, files, args.auto_file)
    out_dir = Path(args.out) if args.out else _default_out_dir(args.flow_id)
    payload = _run_flow(
        flow_id=args.flow_id,
        files=files,
        params=params,
        out_dir=out_dir,
        output_format=args.format,
    )
    if args.json:
        _print_json(payload)
    else:
        print(f"Klart: {args.flow_id}")
        print(f"Output: {out_dir}")
    return 0


def _run_generic(args: argparse.Namespace) -> int:
    flow = _flow_by_id(args.flow_id)
    files = {key: Path(value) for key, value in _parse_key_values(args.file, "file").items()}
    _apply_auto_files(flow, files, args.auto_file)
    params = _parse_key_values(args.param, "param")
    out_dir = Path(args.out) if args.out else _default_out_dir(args.flow_id)
    payload = _run_flow(
        flow_id=args.flow_id,
        files=files,
        params=params,
        out_dir=out_dir,
        output_format=args.format,
    )
    _print_json(payload) if args.json else print(f"Output: {out_dir}")
    return 0


def _run_scenario(args: argparse.Namespace) -> int:
    scenario_path = Path(args.scenario)
    flow_id, files, params, scenario_format = _load_scenario(scenario_path)
    out_dir = Path(args.out) if args.out else _default_out_dir(flow_id)
    payload = _run_flow(
        flow_id=flow_id,
        files=files,
        params=params,
        out_dir=out_dir,
        output_format=args.format or scenario_format or "csv",
    )
    _print_json(payload) if args.json else print(f"Output: {out_dir}")
    return 0


def _run_validate_scenario(args: argparse.Namespace) -> int:
    flow_id, files, params, _scenario_format = _load_scenario(Path(args.scenario))
    flow = _flow_by_id(flow_id)
    errors = _validate_payload(flow, files, params)
    if errors:
        for error in errors:
            print(error, file=sys.stderr)
        return 1
    print("OK")
    return 0


def _run_list_flows(args: argparse.Namespace) -> int:
    public = [_public_flow(flow) for flow in flows.FLOWS]
    if args.format == "json":
        _print_json(public)
        return 0
    headers = ("id", "label", "category")
    print(" | ".join(headers))
    print(" | ".join("---" for _ in headers))
    for flow in public:
        print(f"{flow.get('id')} | {flow.get('label')} | {flow.get('category')}")
    return 0


def _run_schema(args: argparse.Namespace) -> int:
    if args.flow_id == "all":
        _print_json([_public_flow(flow) for flow in flows.FLOWS])
    else:
        _print_json(_public_flow(_flow_by_id(args.flow_id)))
    return 0


def _run_detect(args: argparse.Namespace) -> int:
    path = Path(args.path)
    if not path.exists():
        raise SystemExit(f"Filen finns inte: {path}")
    payload = {"path": str(path), "file_type": detect.detect_file_type(path)}
    _print_json(payload) if args.json else print(payload["file_type"] or "")
    return 0


def _add_output_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--out", help="Outputmapp. Default ar warehouse-cli-out/<flow>-timestamp.")
    parser.add_argument("--format", choices=("csv", "xlsx", "both"), default="csv", help="Resultatexportformat.")
    parser.add_argument("--json", action="store_true", help="Skriv summary till stdout som JSON.")


def _add_auto_file_arg(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--auto-file",
        action="append",
        help="Fil som matchas automatiskt mot flodets inputs via samma detektor som UI:t.",
    )


def _flag_for_key(key: str) -> str:
    return "--" + key.replace("_", "-")


def _add_dynamic_flow_parser(subparsers: argparse._SubParsersAction, flow: dict) -> None:
    parser = subparsers.add_parser(str(flow["id"]), help=str(flow.get("description") or flow.get("label") or ""))
    parser.set_defaults(func=_run_dynamic_command, flow_id=str(flow["id"]))
    for input_def in flow.get("inputs") or []:
        key = str(input_def["key"])
        input_type = input_def.get("type")
        kwargs: dict[str, Any] = {
            "dest": f"input_{key}",
            "required": bool(input_def.get("required")) and input_type != "file",
            "help": str(input_def.get("label") or key),
        }
        if input_type == "number":
            kwargs["type"] = str
        parser.add_argument(_flag_for_key(key), **kwargs)
    _add_auto_file_arg(parser)
    _add_output_args(parser)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    list_parser = subparsers.add_parser("list-flows", help="Lista alla CLI-korbara lagerfloden.")
    list_parser.add_argument("--format", choices=("table", "json"), default="table")
    list_parser.set_defaults(func=_run_list_flows)

    schema_parser = subparsers.add_parser("schema", help="Visa flodesschema.")
    schema_parser.add_argument("flow_id", nargs="?", default="all")
    schema_parser.set_defaults(func=_run_schema)

    detect_parser = subparsers.add_parser("detect", help="Identifiera filtyp.")
    detect_parser.add_argument("path")
    detect_parser.add_argument("--json", action="store_true")
    detect_parser.set_defaults(func=_run_detect)

    run_parser = subparsers.add_parser("run", help="Kor valfritt flode med key=value-argument.")
    run_parser.add_argument("flow_id")
    run_parser.add_argument("--file", action="append", help="Filinput som key=path. Kan anges flera ganger.")
    _add_auto_file_arg(run_parser)
    run_parser.add_argument("--param", action="append", help="Parameter som key=value. Kan anges flera ganger.")
    _add_output_args(run_parser)
    run_parser.set_defaults(func=_run_generic)

    scenario_parser = subparsers.add_parser("run-scenario", help="Kor scenario-json.")
    scenario_parser.add_argument("scenario")
    scenario_parser.add_argument("--out")
    scenario_parser.add_argument("--format", choices=("csv", "xlsx", "both"))
    scenario_parser.add_argument("--json", action="store_true")
    scenario_parser.set_defaults(func=_run_scenario)

    validate_parser = subparsers.add_parser("validate-scenario", help="Validera scenario-json utan att kora.")
    validate_parser.add_argument("scenario")
    validate_parser.set_defaults(func=_run_validate_scenario)

    for flow in flows.FLOWS:
        _add_dynamic_flow_parser(subparsers, flow)

    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    return int(args.func(args) or 0)


if __name__ == "__main__":
    raise SystemExit(main())

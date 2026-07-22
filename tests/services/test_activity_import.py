import io
import asyncio

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.backend.database import Base
from app.backend.models import Activity, Area, Person, ScheduleCell, User
from app.backend.routers.activities import (
    activity_kpi_process_options,
    build_activity_import_template_excel,
    create_activity,
    delete_activity,
    download_import_template,
    import_activities,
    import_activity_rows,
    list_activities,
    parse_activity_import_excel,
    update_activity,
)
from app.backend.schemas import ActivityCreate, ActivityImportRowInput, ActivityImportRowsRequest, ActivityUpdate


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class FakeUpload:
    def __init__(self, content: bytes):
        self.content = content

    async def read(self) -> bytes:
        return self.content


@pytest.fixture()
def import_db():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)


def seed_activity_import_base(db):
    gg = Area(code="GG", name="Gastro Grönt", sort_order=1)
    mg = Area(code="MG", name="Mästergruppen", sort_order=2)
    summary = Activity(code="LEDIGT", label="Ledigt", color="#fee2e2", category="absence", sort_order=10)
    admin = User(username="admin", display_name="Admin", role="admin", roles=["admin"], is_active=True)
    staffing = User(
        username="staffing",
        display_name="Bemanningsansvarig",
        role="staffing_manager",
        roles=["staffing_manager"],
        is_active=True,
    )
    db.add_all([gg, mg, summary, admin, staffing])
    db.flush()
    return gg, mg, summary, admin, staffing


def test_build_activity_import_template_excel_has_expected_headers():
    workbook = load_workbook(io.BytesIO(build_activity_import_template_excel()))
    sheet = workbook.active

    assert [sheet.cell(1, column).value for column in range(1, 8)] == [
        "verksamhet (frivillig)",
        "etikett (obligatorisk)",
        "område (frivillig)",
        "summeras som (frivillig)",
        "KPI Mål (frivillig)",
        "arbetstyp (frivillig)",
        "sortering (frivillig)",
    ]


def test_parse_activity_import_excel_accepts_label_only():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "område", "summeras som", "sortering"],
                ["GG Påfyllning", None, None, None],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].label == "GG Påfyllning"
    assert rows[0].area is None
    assert rows[0].summary_activity is None
    assert rows[0].kpi_process_name is None
    assert rows[0].work_type == "normal"
    assert rows[0].sort_order is None


def test_parse_activity_import_excel_accepts_optional_fields():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "område", "summeras som", "KPI Mål", "arbetstyp", "sortering"],
                ["Frånvaro", "GG", "Ledigt", "dekant ,  plock", "VAS", 20],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].area == "GG"
    assert rows[0].summary_activity == "Ledigt"
    assert rows[0].kpi_process_name == "dekant, plock"
    assert rows[0].work_type == "vas"
    assert rows[0].sort_order == 20


def test_parse_activity_import_excel_rejects_business_prefixed_kpi_process():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "KPI Mål"],
                ["Plock", "GG:decanting, plock"],
            ]
        )
    )

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 2
    assert errors[0].label == "Plock"
    assert errors[0].error == "KPI Mål ska bara vara processnamn, utan bolag"


def test_parse_activity_import_excel_rejects_unknown_work_type():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "arbetstyp"],
                ["Plock", "extra"],
            ]
        )
    )

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 2
    assert errors[0].label == "Plock"
    assert errors[0].error == "Arbetstyp måste vara normal eller VAS"


def test_parse_activity_import_excel_ignores_legacy_category_and_color_columns():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "kategori", "färg", "sortering"],
                ["Gammal mall", "frånvaro", "gul", 7],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].label == "Gammal mall"
    assert rows[0].sort_order == 7


def test_parse_activity_import_excel_collects_row_errors():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "sortering"],
                [None, 1],
                ["Fel sort", "1,5"],
            ]
        )
    )

    assert rows == []
    assert [error.row for error in errors] == [2, 3]
    assert "Etikett" in errors[0].error
    assert "heltal" in errors[1].error


def test_parse_activity_import_excel_requires_label_header():
    with pytest.raises(HTTPException) as exc:
        parse_activity_import_excel(workbook_bytes([["område"], ["GG"]]))

    assert exc.value.status_code == 400


def test_downloaded_activity_template_imports_mixed_optional_summary_and_sorting(import_db):
    _gg, _mg, summary, admin, _staffing = seed_activity_import_base(import_db)
    response = download_import_template(_admin=admin)

    assert response.headers["Content-Disposition"] == 'attachment; filename="aktiviteter-importmall.xlsx"'
    workbook = load_workbook(io.BytesIO(response.body))
    assert workbook.active.title == "Aktiviteter"
    sheet = workbook.active
    assert [sheet.cell(1, column).value for column in range(1, 8)] == [
        "verksamhet (frivillig)",
        "etikett (obligatorisk)",
        "område (frivillig)",
        "summeras som (frivillig)",
        "KPI Mål (frivillig)",
        "arbetstyp (frivillig)",
        "sortering (frivillig)",
    ]

    sheet.append([None, "Test utan frivilligt", "GG", None, None, None, None])
    sheet.append([None, "Test med allt", "MG", "Ledigt", "dekant, plock", "VAS", 42])
    sheet.append([None, "Test bara summeras", None, "Ledigt", None, None, None])
    sheet.append([None, "Test bara sortering", "GG", None, None, None, 77])
    stream = io.BytesIO()
    workbook.save(stream)

    result = asyncio.run(import_activities(file=FakeUpload(stream.getvalue()), db=import_db, admin=admin))

    assert result.created == 4
    assert result.skipped == 0
    assert result.errors == []

    imported = {
        activity.label: activity
        for activity in import_db.query(Activity).filter(Activity.label.like("Test %")).all()
    }
    assert set(imported) == {
        "Test utan frivilligt",
        "Test med allt",
        "Test bara summeras",
        "Test bara sortering",
    }
    assert imported["Test utan frivilligt"].summary_activity_id is None
    assert imported["Test utan frivilligt"].kpi_process_name is None
    assert imported["Test utan frivilligt"].sort_order == 11
    assert imported["Test med allt"].summary_activity_id == summary.id
    assert imported["Test med allt"].kpi_process_name == "dekant, plock"
    assert imported["Test med allt"].work_type == "vas"
    assert imported["Test med allt"].sort_order == 42
    assert imported["Test bara summeras"].summary_activity_id == summary.id
    assert imported["Test bara summeras"].kpi_process_name is None
    assert imported["Test bara summeras"].sort_order == 12
    assert imported["Test bara sortering"].summary_activity_id is None
    assert imported["Test bara sortering"].kpi_process_name is None
    assert imported["Test bara sortering"].sort_order == 77

    for activity in imported.values():
        assert activity.category == "work"
        assert activity.color == "#ffffff"
    assert imported["Test utan frivilligt"].work_type == "normal"


def test_import_activity_rows_creates_from_direct_table(import_db):
    gg, _mg, summary, admin, _staffing = seed_activity_import_base(import_db)

    result = import_activity_rows(
        ActivityImportRowsRequest(
            rows=[
                ActivityImportRowInput(
                    label="Direkt plock",
                    area="GG",
                    summary_activity="Ledigt",
                    kpi_process_name="manual_pick, pack",
                    work_type="VAS",
                    sort_order="31",
                ),
                ActivityImportRowInput(label="Direkt plock", area="GG"),
            ]
        ),
        db=import_db,
        admin=admin,
    )

    assert result.created == 1
    assert result.skipped == 1
    assert result.errors[0].row == 2
    assert result.errors[0].error == "Dubblett i tabellen"
    activity = import_db.query(Activity).filter(Activity.label == "Direkt plock").one()
    assert activity.area_id == gg.id
    assert activity.summary_activity_id == summary.id
    assert activity.kpi_process_name == "manual_pick, pack"
    assert activity.work_type == "vas"
    assert activity.sort_order == 31
    assert activity.category == "work"


def test_bemanningsansvarig_can_manage_activities(import_db):
    gg, _mg, _summary, _admin, staffing = seed_activity_import_base(import_db)

    response = download_import_template(_admin=staffing)
    assert response.headers["Content-Disposition"] == 'attachment; filename="aktiviteter-importmall.xlsx"'

    created = create_activity(
        payload=ActivityCreate(
            label="flow test",
            area_id=gg.id,
            kpi_process_name="dekant ,  plock",
            work_type="VAS",
            sort_order=99,
        ),
        db=import_db,
        admin=staffing,
    )

    assert created.id is not None
    assert created.label == "flow test"
    assert created.kpi_process_name == "dekant, plock"
    assert created.work_type == "vas"
    assert created.category == "work"
    assert created.code.startswith("GG_FLOW_TEST")

    updated = update_activity(
        activity_id=created.id,
        payload=ActivityUpdate(label="flow test uppdaterad", kpi_process_name="manual_pick, pack", work_type="normal"),
        db=import_db,
        admin=staffing,
    )

    assert updated.label == "flow test uppdaterad"
    assert updated.kpi_process_name == "manual_pick, pack"
    assert updated.work_type == "normal"

    absence = update_activity(
        activity_id=created.id,
        payload=ActivityUpdate(category="absence", work_type="vas"),
        db=import_db,
        admin=staffing,
    )

    assert absence.category == "absence"
    assert absence.work_type == "normal"

    cleared = update_activity(
        activity_id=created.id,
        payload=ActivityUpdate(kpi_process_name=""),
        db=import_db,
        admin=staffing,
    )

    assert cleared.kpi_process_name is None

    with pytest.raises(HTTPException) as exc:
        create_activity(
            payload=ActivityCreate(label="fel kpi", area_id=gg.id, kpi_process_name="GG:decanting"),
            db=import_db,
            admin=staffing,
        )

    assert exc.value.status_code == 400
    assert exc.value.detail == "KPI Mål ska bara vara processnamn, utan bolag"

    delete_activity(activity_id=created.id, db=import_db, admin=staffing)

    assert import_db.get(Activity, created.id) is None


def test_activity_kpi_process_options_include_known_rules_and_activity_values(import_db):
    gg, _mg, _summary, _admin, staffing = seed_activity_import_base(import_db)
    activity = Activity(
        code="GG_SPECIAL",
        label="Special",
        area_id=gg.id,
        color="#ffffff",
        category="work",
        sort_order=50,
        is_active=True,
        kpi_process_name="Custom_Process, Manual_Pick",
    )
    import_db.add(activity)
    import_db.commit()

    options = activity_kpi_process_options(db=import_db, user=staffing)
    values = [option.value for option in options]

    assert "Custom_Process" in values
    assert "Manual_Pick" in values
    assert values.count("Manual_Pick") == 1
    assert values == sorted(values, key=str.upper)


def test_activity_delete_removes_inactive_legacy_activity_and_clears_references(import_db):
    gg, _mg, summary, _admin, staffing = seed_activity_import_base(import_db)
    legacy = Activity(
        code="GG_GAMMAL",
        label="Gammal aktivitet",
        area_id=gg.id,
        summary_activity_id=summary.id,
        color="#ffffff",
        category="work",
        sort_order=99,
        is_active=False,
    )
    child = Activity(
        code="GG_BARN",
        label="Barnaktivitet",
        area_id=gg.id,
        summary_activity_id=None,
        color="#ffffff",
        category="work",
        sort_order=100,
        is_active=True,
    )
    person = Person(name="Test Person", home_area_id=gg.id, competencies=[], home_activity_id=None)
    import_db.add_all([legacy, child, person])
    import_db.flush()
    child.summary_activity_id = legacy.id
    person.home_activity_id = legacy.id
    cell = ScheduleCell(
        year=2026,
        week=21,
        weekday=1,
        hour=7,
        minute_start=0,
        minute_end=60,
        person_id=person.id,
        activity_id=legacy.id,
    )
    import_db.add(cell)
    import_db.commit()

    labels = [activity.label for activity in list_activities(include_inactive=False, db=import_db)]
    all_labels = [activity.label for activity in list_activities(include_inactive=True, db=import_db)]
    assert "Gammal aktivitet" not in labels
    assert "Gammal aktivitet" in all_labels

    delete_activity(activity_id=legacy.id, db=import_db, admin=staffing)

    # Aktiviteten har en historisk schemacell: historiken ar en logg, sa
    # aktiviteten inaktiveras (raden behalls for etikett/farg) och den
    # historiska cellen behaller sin aktivitet. Referenser for framtida
    # projektion (hemaktivitet, summeringspekare) rensas fortfarande.
    kept = import_db.get(Activity, legacy.id)
    assert kept is not None
    assert kept.is_active is False
    assert import_db.get(Person, person.id).home_activity_id is None
    assert import_db.get(Activity, child.id).summary_activity_id is None
    historical_cell = import_db.get(ScheduleCell, cell.id)
    assert historical_cell.activity_id == legacy.id
    assert historical_cell.empty_override is False

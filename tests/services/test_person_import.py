import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.backend.database import Base
from app.backend.models import Activity, Area, Person, User
from app.backend.routers.persons import (
    build_person_import_template_excel,
    create_person,
    delete_person,
    import_person_rows,
    parse_person_import_excel,
    update_person,
)
from app.backend.schemas import PersonCreate, PersonImportRowInput, PersonImportRowsRequest, PersonUpdate


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_build_person_import_template_excel_has_expected_headers():
    workbook = load_workbook(io.BytesIO(build_person_import_template_excel()))
    sheet = workbook.active

    assert [sheet.cell(1, column).value for column in range(1, 9)] == [
        "verksamhet (frivillig)",
        "namn (obligatorisk)",
        "NoMan (obligatorisk)",
        "RFID (frivillig)",
        "arbetstyp (frivillig)",
        "hemomr\u00e5de (frivillig)",
        "huvudaktivitet (frivillig)",
        "sortering (frivillig)",
    ]


def test_parse_person_import_excel_requires_noman():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn", "hemomr\u00e5de", "huvudaktivitet", "sortering"],
                ["Anna Andersson", None, None, None],
            ]
        )
    )

    assert rows == []
    assert len(errors) == 1
    assert errors[0].name == "Anna Andersson"
    assert errors[0].error == "NoMan saknas"


def test_parse_person_import_excel_accepts_optional_fields():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn (obligatorisk)", "NoMan (obligatorisk)", "RFID", "hemomr\u00e5de (frivillig)", "huvudst\u00e4lle (frivillig)", "sortering (frivillig)"],
                ["Bo Berg", "BOB01", "rfid-77", "GG", "GG VM", 12],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].noman == "BOB01"
    assert rows[0].rfid_code == "RFID-77"
    assert rows[0].collar_type == "blue_collar"
    assert rows[0].home_area == "GG"
    assert rows[0].home_activity == "GG VM"
    assert rows[0].sort_order == 12


def test_parse_person_import_excel_accepts_collar_type_aliases():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn", "NoMan", "arbetstyp"],
                ["Anna Andersson", "ANN01", "White color"],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].collar_type == "white_collar"


def test_parse_person_import_excel_collects_row_errors():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn", "NoMan", "sortering"],
                [None, "ROW02", 1],
                ["Cecilia", "CEC01", "1,5"],
            ]
        )
    )

    assert rows == []
    assert [error.row for error in errors] == [2, 3]
    assert "Namn" in errors[0].error
    assert "heltal" in errors[1].error


def test_parse_person_import_excel_requires_name_header():
    with pytest.raises(HTTPException) as exc:
        parse_person_import_excel(workbook_bytes([["hemomr\u00e5de"], ["GG"]]))

    assert exc.value.status_code == 400


@pytest.fixture()
def person_db():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)


def test_create_person_rejects_duplicate_name(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    person_db.add_all([admin, Person(name="Anna Andersson", competencies=[], is_active=True, sort_order=1)])
    person_db.flush()

    with pytest.raises(HTTPException) as exc:
        create_person(PersonCreate(name=" anna andersson ", noman="ANN01"), db=person_db, user=admin)

    assert exc.value.status_code == 409


def test_import_person_rows_creates_from_direct_table(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    area = Area(code="GG", name="Granngården", sort_order=1)
    person_db.add_all([admin, area])
    person_db.flush()
    activity = Activity(
        code="GG_PLOCK",
        label="GG Plock",
        area_id=area.id,
        color="#bfdbfe",
        category="work",
        sort_order=2,
        is_active=True,
    )
    person_db.add(activity)
    person_db.flush()

    result = import_person_rows(
        PersonImportRowsRequest(
            rows=[
                PersonImportRowInput(name="Mira Multi", noman="MIR01", rfid_code="abc123", collar_type="white_collar", home_area="GG", home_activity="GG Plock", sort_order="7"),
                PersonImportRowInput(name="Mira Multi", noman="MIR02", home_area="GG"),
            ]
        ),
        db=person_db,
        user=admin,
    )

    assert result.created == 1
    assert result.skipped == 1
    assert result.errors[0].row == 2
    assert result.errors[0].error == "Dubblett i tabellen"
    person = person_db.query(Person).filter(Person.name == "Mira Multi").one()
    assert person.home_area_id == area.id
    assert person.home_activity_id == activity.id
    assert person.noman == "MIR01"
    assert person.rfid_code == "ABC123"
    assert person.collar_type == "white_collar"
    assert person.sort_order == 7


def test_import_person_rows_requires_noman(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    person_db.add(admin)
    person_db.flush()

    result = import_person_rows(
        PersonImportRowsRequest(rows=[PersonImportRowInput(name="Mira Multi")]),
        db=person_db,
        user=admin,
    )

    assert result.created == 0
    assert result.skipped == 1
    assert result.errors[0].name == "Mira Multi"
    assert result.errors[0].error == "NoMan saknas"


def test_create_person_requires_noman(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    person_db.add(admin)
    person_db.flush()

    with pytest.raises(HTTPException) as exc:
        create_person(PersonCreate(name="NoMan Person"), db=person_db, user=admin)

    assert exc.value.status_code == 400
    assert exc.value.detail == "NoMan krävs"


def test_create_and_update_person_persists_required_noman(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    person_db.add(admin)
    person_db.flush()

    person = create_person(PersonCreate(name="NoMan Person", noman="NMP01"), db=person_db, user=admin)
    assert person.noman == "NMP01"
    assert person.collar_type == "blue_collar"

    updated = update_person(person.id, PersonUpdate(noman="NMP02", collar_type="white_collar"), db=person_db, user=admin)

    assert updated.noman == "NMP02"
    assert updated.collar_type == "white_collar"

    with pytest.raises(HTTPException) as exc:
        update_person(person.id, PersonUpdate(noman=None), db=person_db, user=admin)

    assert exc.value.status_code == 400
    assert exc.value.detail == "NoMan krävs"


def test_create_update_and_import_reject_duplicate_rfid(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    person_db.add(admin)
    person_db.flush()

    first = create_person(PersonCreate(name="RFID One", noman="RF101", rfid_code="tag-1"), db=person_db, user=admin)
    second = create_person(PersonCreate(name="RFID Two", noman="RF102"), db=person_db, user=admin)

    assert first.rfid_code == "TAG-1"

    with pytest.raises(HTTPException) as exc:
        update_person(second.id, PersonUpdate(rfid_code=" tag-1 "), db=person_db, user=admin)
    assert exc.value.status_code == 409

    result = import_person_rows(
        PersonImportRowsRequest(rows=[PersonImportRowInput(name="RFID Three", noman="RF103", rfid_code="TAG-1")]),
        db=person_db,
        user=admin,
    )

    assert result.created == 0
    assert result.skipped == 1
    assert result.errors[0].error == "RFID-kod finns redan"


def test_create_person_rejects_inactive_duplicate_name(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    inactive = Person(name="Anton Holmqvist", competencies=[], is_active=False, sort_order=17)
    person_db.add_all([admin, inactive])
    person_db.flush()

    with pytest.raises(HTTPException) as exc:
        create_person(PersonCreate(name="Anton Holmqvist", noman="ANT01", sort_order=3), db=person_db, user=admin)

    assert exc.value.status_code == 409
    assert inactive.is_active is False
    assert person_db.query(Person).filter(Person.name == "Anton Holmqvist").count() == 1


def test_delete_person_removes_person(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    person = Person(name="Bo Berg", competencies=[], is_active=False, sort_order=17)
    person_db.add_all([admin, person])
    person_db.flush()

    delete_person(person.id, db=person_db, user=admin)

    assert person_db.get(Person, person.id) is None


def test_update_person_rejects_duplicate_name(person_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    anna = Person(name="Anna Andersson", competencies=[], is_active=True, sort_order=1)
    bo = Person(name="Bo Berg", competencies=[], is_active=True, sort_order=2)
    person_db.add_all([admin, anna, bo])
    person_db.flush()

    with pytest.raises(HTTPException) as exc:
        update_person(bo.id, PersonUpdate(name="Anna Andersson"), db=person_db, user=admin)

    assert exc.value.status_code == 409

import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.backend.database import Base
from app.backend.models import Area, User
from app.backend.routers.users import build_user_import_template_excel, import_user_rows, parse_user_import_excel
from app.backend.schemas import UserImportRowInput, UserImportRowsRequest


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_build_user_import_template_excel_has_expected_headers():
    workbook = load_workbook(io.BytesIO(build_user_import_template_excel()))
    sheet = workbook.active

    assert [sheet.cell(1, column).value for column in range(1, 6)] == [
        "verksamhet (frivillig)",
        "anv\u00e4ndarnamn (obligatorisk)",
        "namn (frivillig)",
        "roller (obligatorisk)",
        "omr\u00e5de (frivillig)",
    ]


def test_parse_user_import_excel_accepts_template_headers_and_swedish_roles():
    content = workbook_bytes(
        [
            ["anv\u00e4ndarnamn (obligatorisk)", "namn (frivillig)", "roll (obligatorisk)", "omr\u00e5de (frivillig)"],
            ["anna", "Anna Andersson", "arbetsledare", ""],
            ["petra", "Petra flow", "bemanningsansvarig", ""],
            ["bo", "Bo Berg", "administrat\u00f6r", "GG"],
            ["viola", "Viola Visning", "visning", "Mestergruppen"],
            ["lina", "Lina Lager", "lagerkontorist", ""],
            ["arvid", "Arvid Artikel", "artikelplacerare", ""],
        ]
    )

    rows, errors = parse_user_import_excel(content)

    assert errors == []
    assert [(row.username, row.display_name, row.roles, row.area_name) for row in rows] == [
        ("anna", "Anna Andersson", ["leader"], None),
        ("petra", "Petra flow", ["staffing_manager"], None),
        ("bo", "Bo Berg", ["admin"], "GG"),
        ("viola", "Viola Visning", ["viewer"], "Mestergruppen"),
        ("lina", "Lina Lager", ["warehouse_clerk"], None),
        ("arvid", "Arvid Artikel", ["article_placer"], None),
    ]


def test_parse_user_import_excel_accepts_multiple_roles_in_one_cell():
    content = workbook_bytes(
        [
            ["username", "name", "roles"],
            ["mira", "Mira Multi", "admin, arbetsledare"],
        ]
    )

    rows, errors = parse_user_import_excel(content)

    assert errors == []
    assert rows[0].roles == ["admin", "leader"]


def test_parse_user_import_excel_collects_row_errors():
    content = workbook_bytes(
        [
            ["anv\u00e4ndarnamn", "namn", "roll"],
            [None, "Namnl\u00f6s", "arbetsledare"],
            ["cecilia", "Cecilia", "ok\u00e4nd"],
        ]
    )

    rows, errors = parse_user_import_excel(content)

    assert rows == []
    assert [error.row for error in errors] == [2, 3]
    assert "Anv\u00e4ndarnamn" in errors[0].error
    assert "Roll" in errors[1].error


def test_parse_user_import_excel_requires_expected_headers():
    with pytest.raises(HTTPException) as exc:
        parse_user_import_excel(workbook_bytes([["username", "name"], ["anna", "Anna"]]))

    assert exc.value.status_code == 400


@pytest.fixture()
def user_db():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)


def test_import_user_rows_creates_from_direct_table(user_db):
    admin = User(username="admin", role="admin", roles=["admin"], is_active=True)
    area = Area(code="GG", name="Granngården", sort_order=1)
    user_db.add_all([admin, area])
    user_db.flush()

    result = import_user_rows(
        UserImportRowsRequest(
            rows=[
                UserImportRowInput(
                    username="mira",
                    display_name="Mira Multi",
                    roles="admin, arbetsledare",
                    area="GG",
                ),
                UserImportRowInput(username="mira", display_name="Dubblett", roles="visning"),
            ]
        ),
        db=user_db,
        admin=admin,
    )

    assert result.created == 1
    assert result.skipped == 1
    assert result.errors[0].row == 2
    assert result.errors[0].error == "Dubblett i tabellen"
    imported = user_db.query(User).filter(User.username == "mira").one()
    assert imported.display_name == "Mira Multi"
    assert imported.roles == ["admin", "leader"]
    assert imported.area_id == area.id
    assert imported.must_change_password is True
